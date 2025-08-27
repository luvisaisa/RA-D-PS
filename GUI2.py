"""Simple GUI application for parsing XML files and exporting results."""

from __future__ import annotations

import os
import re
import tkinter as tk
from tkinter import filedialog, messagebox
import tkinter.font as tkfont

from XMLPARSE import NYTXMLParser


class NYTXMLGuiApp:
    """GUI application to parse XML files and optionally export to Excel."""

    def __init__(self, master: tk.Tk) -> None:
        self.master = master
        self.master.title("NYT XML Parser")
        self.parser = NYTXMLParser()
        self.files: list[str] = []
        self._create_widgets()

    def _create_widgets(self) -> None:
        frame = tk.Frame(self.master)
        frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        try:
            font = tkfont.Font(family="Aptos", size=11)
        except tk.TclError:
            font = tkfont.Font(family="Arial", size=11)

        # set your preferred color for the file view box
        listbox_bg = "#313d5c"  # a dark blue

        self.listbox = tk.Listbox(
            frame,
            height=6,
            font=font,
            bg="#bbcfff",                # light blue background
            fg="#0a0e3c",                # dark blue text when not selected
            selectmode=tk.EXTENDED,
            selectbackground="#0044bb",  # bright blue highlight when selected
            selectforeground="#ffffff",  # white text when selected
            highlightthickness=2,        # visible border
            highlightbackground="#4f6bed", # border color
            bd=2,
            relief=tk.GROOVE
        )
        self.listbox.pack(fill=tk.BOTH, expand=True, pady=5)

        btn_files = tk.Button(
            frame, text="Select XML Files", command=self.select_files, font=tkfont.Font(family="Aptos", size=11), bg="#d7e3fc",
            highlightthickness=0, bd=0, relief=tk.FLAT
        )
        btn_files.pack(fill=tk.X)

        btn_folder = tk.Button(
            frame, text="Select Folder", command=self.select_folder, font=tkfont.Font(family="Aptos", size=11), bg="#d7e3fc",
            highlightthickness=0, bd=0, relief=tk.FLAT
        )
        btn_folder.pack(fill=tk.X, pady=(5, 0))

        btn_excel = tk.Button(
            frame, text="Select Excel to Append", command=self.select_excel, font=tkfont.Font(family="Aptos", size=11), bg="#d7e3fc",
            highlightthickness=0, bd=0, relief=tk.FLAT
        )
        btn_excel.pack(fill=tk.X)

        # create a frame for export and clear buttons side by side
        export_frame = tk.Frame(frame)
        export_frame.pack(fill=tk.X, pady=(5, 0))

        btn_export_new = tk.Button(
            export_frame,
            text="Select New Excel File",
            command=lambda: self._export_to_excel([]),  # or provide actual data if needed
            font=tkfont.Font(family="Aptos", size=11),
            bg="#228B22",
            fg="#ffffff",
            highlightthickness=0,
            bd=0,
            relief=tk.FLAT
        )
        btn_export_new.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 2))

        clear_btn = tk.Button(
            export_frame,
            text="Clear Files",
            command=self.clear_files,
            font=font,
            bg="#d7e3fc",
            highlightthickness=0,
            bd=0,
            relief=tk.FLAT
        )
        clear_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(2, 0))

        parse_btn = tk.Button(
            frame, text="Append to Selected Excel", command=self.parse_files, font=tkfont.Font(family="Aptos", size=11), bg="#d7e3fc",
            highlightthickness=0, bd=0, relief=tk.FLAT
        )
        parse_btn.pack(fill=tk.X)

    def select_files(self) -> None:
        filenames = filedialog.askopenfilenames(
            title="Choose XML files", filetypes=[("XML Files", "*.xml")]
        )
        if filenames:
            self.files = sorted(list(filenames))
            self._update_file_list()

    def select_folder(self) -> None:
        folder = filedialog.askdirectory(title="Choose folder with XML files")
        if folder:
            all_files = [
                os.path.join(folder, f)
                for f in os.listdir(folder)
                if os.path.isfile(os.path.join(folder, f))
            ]
            xml_files = [f for f in all_files if f.lower().endswith(".xml")]
            if not xml_files:
                if all_files:
                    self.show_temporary_error("No XML files were found in the selected folder.\nOther file types are present.")
                else:
                    self.show_temporary_error("The selected folder is empty.")
                return
            self.files = sorted(xml_files, key=natural_sort_key)
            print([os.path.basename(f) for f in self.files])  # <-- Add here for debugging
            self._update_file_list()

    def _update_file_list(self) -> None:
        self.listbox.delete(0, tk.END)
        for path in self.files:
            self.listbox.insert(tk.END, os.path.basename(path))
        self.listbox.select_set(0, tk.END)
        print([os.path.basename(f) for f in self.files])  # Debugging line

    def parse_files(self) -> None:
        if not self.files:
            messagebox.showinfo("No files", "Please select XML files to parse.")
            return
        parsed = []
        for path in self.files:
            try:
                with open(path, "r", encoding="utf-8") as fh:
                    content = fh.read()
                root = self.parser.parse(content)
                parsed.append((os.path.basename(path), root.tag))
            except Exception as exc:  # pragma: no cover - GUI message
                messagebox.showerror("Parse error", f"Failed to parse {path}: {exc}")
                return
        self._ask_export(parsed)

    def _ask_export(self, data: list[tuple[str, str]]) -> None:
        if messagebox.askyesno("Export", "Export parsed data to Excel?"):
            self._export_to_excel(data)

    def _export_to_excel(self, data: list[tuple[str, str]]) -> None:
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError:
            messagebox.showerror(
                "Missing dependency", "openpyxl is required to export to Excel."
            )
            return

        path = filedialog.asksaveasfilename(
            title="Save Excel file",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
        )
        if not path:
            return

        if os.path.exists(path):
            wb = load_workbook(path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Filename", "RootTag"])

        for filename, tag in data:
            ws.append([filename, tag])

        wb.save(path)
        messagebox.showinfo("Exported", f"Data exported to {path}")

    def show_temporary_error(self, message: str):
        self.master.lift()
        popup = tk.Toplevel(self.master)
        popup.overrideredirect(True)
        popup.configure(bg="#8B0000")
        popup.attributes("-topmost", True)

        # Always update to get the latest window size and position
        self.master.update_idletasks()
        width = 240
        height = 40
        # Get current window position and size
        win_x = self.master.winfo_x()
        win_y = self.master.winfo_y()
        win_w = self.master.winfo_width()
        # Center the popup at the top of the window
        x = win_x + (win_w // 2) - (width // 2)
        y_start = win_y - height
        y_end = win_y + 20

        popup.geometry(f"{width}x{height}+{x}+{y_start}")

        label = tk.Label(
            popup, text=message, bg="#8B0000", fg="white",
            font=("Aptos", 11, "bold")
        )
        label.pack(expand=True, fill=tk.BOTH)

        popup.lift()
        popup.focus_force()

        # Animation: slide down
        steps = 10
        duration = 200  # ms for slide down
        delay = duration // steps
        delta = (y_end - y_start) / steps

        def slide_down(step=0):
            new_y = int(y_start + delta * step)
            popup.geometry(f"{width}x{height}+{x}+{new_y}")
            if step < steps:
                popup.after(delay, slide_down, step + 1)
            else:
                popup.after(4000, slide_up, steps)

        def slide_up(step):
            new_y = int(y_start + delta * step)
            popup.geometry(f"{width}x{height}+{x}+{new_y}")
            if step > 0:
                popup.after(delay, slide_up, step - 1)
            else:
                popup.destroy()

        slide_down()

    def clear_files(self) -> None:
        self.files = []
        self._update_file_list()
        self.show_temporary_error("File list cleared.")

    def select_excel(self):
        messagebox.showinfo("Not implemented", "This feature is not implemented yet.")

def natural_sort_key(s):
    return [int(text) if text.isdigit() else text.lower() for text in re.split(r'(\d+)', os.path.basename(s))]


def main() -> None:
    root = tk.Tk()
    NYTXMLGuiApp(root)
    root.mainloop()


if __name__ == "__main__":  # pragma: no cover - manual execution
    main()
