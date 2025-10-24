## draft 2

# import necessary libraries for data handling, xml parsing, os operations, and gui
import datetime
import gc
import os
import pandas as pd
import platform
import re
import subprocess
import traceback
import xml.etree.ElementTree as ET
from collections import defaultdict
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

# Import the database module
try:
    from radiology_database import RadiologyDatabase
    SQLITE_AVAILABLE = True
except ImportError:
    SQLITE_AVAILABLE = False
    print("⚠️ SQLite database features unavailable - install required packages or check radiology_database.py")

def open_file_cross_platform(file_path):
    """Open a file using the default system application across different platforms"""
    try:
        if platform.system() == "Darwin":  # macOS
            subprocess.run(["open", file_path], check=True)
        elif platform.system() == "Windows":  # Windows
            # os.startfile is Windows-only; use subprocess for cross-platform
            try:
                os.startfile(file_path)
            except AttributeError:
                print("os.startfile not available on this platform.")
        else:  # Linux and others
            subprocess.run(["xdg-open", file_path], check=True)
    except Exception as e:
        print(f"Could not open file {file_path}: {e}")

# -------- RA-D-PS Excel Exporter --------

def _sanitize_name(name: str) -> str:
    """Keep A-Z a-z 0-9 _ -, replace others with underscore."""
    return re.sub(r"[^A-Za-z0-9_\-]+", "_", name.strip())

def _timestamp() -> str:
    """Return local timestamp YYYY-MM-DD_HHMMSS."""
    return datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")

def _next_versioned_path(folder_path: str, base_filename: str) -> str:
    """
    If base filename exists in folder, append _v2, _v3, ... until unused.
    Returns full path.
    """
    path = os.path.join(folder_path, base_filename)
    if not os.path.exists(path):
        return path
    base_root, ext = os.path.splitext(base_filename)
    i = 2
    while True:
        candidate = os.path.join(folder_path, f"{base_root}_v{i}{ext}")
        if not os.path.exists(candidate):
            return candidate
        i += 1

def _count_numbered_keys(d, prefix: str) -> int:
    """Count numbered keys like radiologist_1, radiologist_2, etc."""
    nums = []
    for k in d.keys():
        m = re.fullmatch(rf"{re.escape(prefix)}_(\d+)", str(k))
        if m:
            nums.append(int(m.group(1)))
    return max(nums) if nums else 0

def _get_R_max(records: list, force_blocks=None) -> int:
    """Determine maximum radiologist count across records."""
    R_max = 0
    for rec in records:
        cand = 0
        if isinstance(rec.get("radiologist_count"), int):
            cand = max(cand, rec["radiologist_count"])
        if isinstance(rec.get("radiologists"), dict):
            cand = max(cand, len(rec["radiologists"]))
        cand = max(cand, _count_numbered_keys(rec, "radiologist"))
        R_max = max(R_max, cand)
    if force_blocks is not None:
        R_max = max(R_max, int(force_blocks))
    return R_max

def _build_columns(R_max: int) -> list:
    """
    Build the full column plan as a list where None represents a spacer.
    Fixed columns first, then R_max radiologist blocks each followed by a spacer.
    """
    cols = ["file #", "Study UID", None, "NoduleID", None]
    for r in range(1, R_max + 1):
        cols += [
            f"Radiologist_{r}_Subtlety",
            f"Radiologist_{r}_Confidence",
            f"Radiologist_{r}_Obscuration",
            f"Radiologist_{r}_Reason",
            "Coordinates",
            None,  # spacer between radiologists
        ]
    return cols

def _non_spacer_col_indices(cols: list) -> list:
    """Return a list of 1-based column indices that are NOT spacers (i.e., not None)."""
    return [i + 1 for i, c in enumerate(cols) if c is not None]

def _apply_row_striping(ws, non_spacer_indices: list, blue_argb: str = "FFCCE5FF", white_argb: str = "FFFFFFFF"):
    """Apply alternating row striping via conditional formatting (DifferentialStyle) on non-spacer columns."""
    dxf_blue = DifferentialStyle(fill=PatternFill(start_color=blue_argb, end_color=blue_argb, fill_type="solid"))
    dxf_white = DifferentialStyle(fill=PatternFill(start_color=white_argb, end_color=white_argb, fill_type="solid"))
    rule_blue = Rule(type="expression", dxf=dxf_blue, formula=["MOD(ROW(),2)=0"])
    rule_white = Rule(type="expression", dxf=dxf_white, formula=["MOD(ROW(),2)=1"])
    max_rows = 1048576
    for idx in non_spacer_indices:
        col_letter = get_column_letter(idx)
        rng = f"{col_letter}1:{col_letter}{max_rows}"
        ws.conditional_formatting.add(rng, rule_blue)
        ws.conditional_formatting.add(rng, rule_white)

def _auto_size_columns(ws, cols: list):
    """Auto-size all columns based on max cell length; keep spacers narrow."""
    for i, header in enumerate(cols, start=1):
        col_letter = get_column_letter(i)
        if header is None:
            ws.column_dimensions[col_letter].width = 3  # spacer stays thin
            continue
        max_len = 0
        if header:
            max_len = max(max_len, len(str(header)))
        for cell in ws[col_letter]:
            v = cell.value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = max(10, min(max_len + 2, 60))

def _fill_spacer_columns(ws, cols, blue_argb="FFCCE5FF"):
    """Fill all spacer columns with solid blue background."""
    spacer_fill = PatternFill(start_color=blue_argb, end_color=blue_argb, fill_type="solid")
    for i, header in enumerate(cols, start=1):
        if header is None:
            col_letter = get_column_letter(i)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=i, max_col=i):
                for cell in row:
                    cell.fill = spacer_fill

def _set_column_widths(ws, cols: list):
    """Apply reasonable default widths and widen Reason columns."""
    for i, header in enumerate(cols, start=1):
        if header is None:
            ws.column_dimensions[get_column_letter(i)].width = 3  # narrow spacer
        else:
            if str(header).endswith("_Reason"):
                ws.column_dimensions[get_column_letter(i)].width = 24
            else:
                ws.column_dimensions[get_column_letter(i)].width = 12

def export_excel(records, folder_path, sheet="radiology_data", blue_argb="FFCCE5FF", force_blocks=None):
    """
    Export records to Excel with dynamic radiologist blocks, spacer columns (solid blue fill),
    alternating row striping via conditional formatting, and true auto-sizing.
    Returns output_path.
    
    Args:
        records: list of dicts with required keys:
            - file_number (str/int), study_uid (str), nodule_id (str)
            - EITHER radiologist_count (int) OR radiologists dict "1".."R" -> dict(
                  subtlety, confidence, obscuration, reason, coordinates
              )
        folder_path: where to save the file and derive base name
        sheet: worksheet name (default "radiology_data")
        blue_argb: ARGB color for blue stripe and spacers (default FFCCE5FF)
        force_blocks: minimum number of radiologist blocks to create
    """
    if not isinstance(records, list):
        raise ValueError("records must be a list of dictionaries")

    # Ensure folder exists
    os.makedirs(folder_path, exist_ok=True)

    # Determine R_max for header
    R_max = _get_R_max(records, force_blocks=force_blocks)
    cols = _build_columns(R_max)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet

    # Header row
    for i, header in enumerate(cols, start=1):
        if header is not None:
            ws.cell(row=1, column=i, value=header)

    # Data rows
    r_ptr = 2
    for rec in records:
        ws.cell(row=r_ptr, column=1, value=rec.get("file_number"))
        ws.cell(row=r_ptr, column=2, value=rec.get("study_uid"))
        ws.cell(row=r_ptr, column=4, value=rec.get("nodule_id"))

        if isinstance(rec.get("radiologists"), dict) and rec["radiologists"]:
            try:
                r_keys = sorted(rec["radiologists"].keys(), key=lambda x: int(x))
            except Exception:
                r_keys = sorted(rec["radiologists"].keys())
            R_this = len(r_keys)
        else:
            R_this = int(rec.get("radiologist_count", 0))
            if R_this == 0:
                R_this = _count_numbered_keys(rec, "radiologist")
            r_keys = [str(i) for i in range(1, R_this + 1)]

        c_ptr = 6
        for r in range(1, R_max + 1):
            if r <= R_this:
                if "radiologists" in rec and isinstance(rec["radiologists"], dict):
                    rdict = rec["radiologists"].get(str(r), {})
                else:
                    rdict = rec.get(f"radiologist_{r}", {}) or {}
                ws.cell(row=r_ptr, column=c_ptr + 0, value=rdict.get("subtlety"))
                ws.cell(row=r_ptr, column=c_ptr + 1, value=rdict.get("confidence"))
                ws.cell(row=r_ptr, column=c_ptr + 2, value=rdict.get("obscuration"))
                ws.cell(row=r_ptr, column=c_ptr + 3, value=rdict.get("reason"))
                ws.cell(row=r_ptr, column=c_ptr + 4, value=rdict.get("coordinates"))
            c_ptr += 6

        r_ptr += 1

    # conditional formatting: row striping on non-spacer columns
    non_spacers = _non_spacer_col_indices(cols)
    _apply_row_striping(ws, non_spacers, blue_argb=blue_argb, white_argb="FFFFFFFF")

    # freeze panes
    ws.freeze_panes = "A2"

    # auto-size columns
    _auto_size_columns(ws, cols)

    # fill spacer columns solid blue
    _fill_spacer_columns(ws, cols, blue_argb=blue_argb)

    # Auto-naming
    folder_name = os.path.basename(os.path.abspath(folder_path)) or "export"
    folder_name = _sanitize_name(folder_name)
    ts = _timestamp()
    base_filename = f"{folder_name}_RA-D-PS_{ts}.xlsx"
    out_path = _next_versioned_path(folder_path, base_filename)

    # Save
    wb.save(out_path)
    return out_path

# -------- End RA-D-PS Excel Exporter --------

def convert_parsed_data_to_ra_d_ps_format(dataframes):
    """
    Convert parsed XML DataFrames to RA-D-PS format records.
    
    Args:
        dataframes: tuple of (main_df, unblinded_df) from parse_radiology_sample
                   or dict of {parse_case: df} from parse_multiple
    
    Returns:
        list of records suitable for export_excel function
    """
    records = []
    
    # Handle different input formats
    if isinstance(dataframes, tuple):
        # Single file result: (main_df, unblinded_df)
        main_df, unblinded_df = dataframes
        print(f"🔄 Converting RA-D-PS format...")
        print(f"  📊 Main DataFrame: {len(main_df)} rows")
        print(f"  📊 Unblinded DataFrame: {len(unblinded_df)} rows")
        
        combined_df = pd.concat([main_df, unblinded_df], ignore_index=True) if not main_df.empty or not unblinded_df.empty else pd.DataFrame()
        print(f"  📊 Combined DataFrame: {len(combined_df)} rows")
        
        dfs_to_process = [combined_df] if not combined_df.empty else []
    elif isinstance(dataframes, dict):
        # Multiple files result: {parse_case: df}
        print(f"🔄 Converting multiple DataFrames to RA-D-PS format...")
        dfs_to_process = [df for df in dataframes.values() if not df.empty]
        print(f"  📊 Processing {len(dfs_to_process)} DataFrames")
    elif isinstance(dataframes, pd.DataFrame):
        # Single DataFrame
        print(f"🔄 Converting single DataFrame to RA-D-PS format...")
        dfs_to_process = [dataframes] if not dataframes.empty else []
    else:
        print(f"⚠️  Unknown dataframes format: {type(dataframes)}")
        dfs_to_process = []
    
    for df in dfs_to_process:
        if df.empty:
            continue
            
        print(f"  🔍 Processing DataFrame with {len(df)} rows")
        
        # Group by file and nodule to aggregate radiologist data
        grouped = df.groupby(['FileID', 'NoduleID'])
        print(f"  📋 Found {len(grouped)} unique file/nodule combinations")
        
        for (file_id, nodule_id), group in grouped:
            print(f"    📄 Processing {file_id} - {nodule_id} ({len(group)} rows)")
            
            # Extract study UID from first row
            study_uid = group['StudyInstanceUID'].iloc[0] if 'StudyInstanceUID' in group.columns else "N/A"
            
            # Build radiologists dictionary
            print(f"      👥 Building radiologists dictionary...")
            radiologists = {}
            for idx, row in group.iterrows():
                radiologist = row.get('Radiologist', f'rad_{idx+1}')
                print(f"        👨‍⚕️ Processing radiologist: {radiologist}")
                
                # Extract radiologist number from name (e.g., "anonRad1" -> "1")
                rad_num_match = re.search(r'(\d+)', str(radiologist))
                rad_num = rad_num_match.group(1) if rad_num_match else str(len(radiologists) + 1)
                print(f"        🔢 Extracted rad_num: {rad_num}")
                
                # Build coordinates string
                x_coord = row.get('X_coord', '')
                y_coord = row.get('Y_coord', '')
                z_coord = row.get('Z_coord', '')
                coordinates = f"{x_coord}, {y_coord}, {z_coord}" if any([x_coord, y_coord, z_coord]) else ""
                print(f"        📍 Coordinates: {coordinates}")
                
                radiologists[rad_num] = {
                    "subtlety": row.get('Subtlety', ''),
                    "confidence": row.get('Confidence', ''),
                    "obscuration": row.get('Obscuration', ''),
                    "reason": row.get('Reason', ''),
                    "coordinates": coordinates.strip(", ")
                }
                print(f"        ✅ Added rad_num {rad_num} to dictionary")
            
            print(f"      📝 Final radiologists dictionary has {len(radiologists)} entries: {list(radiologists.keys())}")
            
            record = {
                "file_number": file_id,
                "study_uid": study_uid,
                "nodule_id": nodule_id,
                "radiologists": radiologists
            }
            
            print(f"      📋 Created record with {len(radiologists)} radiologists")
            records.append(record)
    
    return records

# function to parse a single radiology xml file
def get_expected_attributes_for_case(parse_case):
    """define what attributes should be expected for each parse case"""
    expected_attrs = {
        "Complete_Attributes": {
            "header": ["StudyInstanceUID", "SeriesInstanceUID", "Modality", "DateService", "TimeService"],
            "characteristics": ["confidence", "subtlety", "obscuration", "reason"],
            "roi": ["imageSOP_UID", "xCoord", "yCoord"],
            "nodule": ["noduleID"]
        },
        "With_Reason_Partial": {
            "header": ["StudyInstanceUID", "SeriesInstanceUID", "Modality", "DateService", "TimeService"],
            "characteristics": ["reason", "confidence", "subtlety"],  # Partial but has reason
            "roi": ["imageSOP_UID", "xCoord", "yCoord"],
            "nodule": ["noduleID"]
        },
        "Core_Attributes_Only": {
            "header": ["StudyInstanceUID", "SeriesInstanceUID", "Modality"],
            "characteristics": ["confidence", "subtlety", "obscuration"],  # No reason
            "roi": ["imageSOP_UID", "xCoord", "yCoord"],
            "nodule": ["noduleID"]
        },
        "Minimal_Attributes": {
            "header": ["StudyInstanceUID", "SeriesInstanceUID"],
            "characteristics": ["confidence"],  # Only one attribute
            "roi": ["imageSOP_UID"],
            "nodule": ["noduleID"]
        },
        "No_Characteristics": {
            "header": ["StudyInstanceUID", "SeriesInstanceUID"],
            "characteristics": [],  # No characteristics expected
            "roi": ["imageSOP_UID"],
            "nodule": ["noduleID"]
        },
        "LIDC_Single_Session": {
            "header": ["StudyInstanceUID", "SeriesInstanceUID", "DateService", "TimeService"],  # Modality often missing
            "characteristics": ["subtlety"],  # Only subtlety commonly present
            "roi": ["imageSOP_UID", "xCoord", "yCoord"],
            "nodule": ["noduleID"]
        },
        "LIDC_Multi_Session_2": {
            "header": ["StudyInstanceUID", "SeriesInstanceUID", "DateService", "TimeService"],  # Modality often missing
            "characteristics": ["subtlety"],  # Only subtlety commonly present
            "roi": ["imageSOP_UID", "xCoord", "yCoord"],
            "nodule": ["noduleID"]
        },
        "LIDC_Multi_Session_3": {
            "header": ["StudyInstanceUID", "SeriesInstanceUID", "DateService", "TimeService"],  # Modality often missing
            "characteristics": ["subtlety"],  # Only subtlety commonly present
            "roi": ["imageSOP_UID", "xCoord", "yCoord"],
            "nodule": ["noduleID"]
        },
        "LIDC_Multi_Session_4": {
            "header": ["StudyInstanceUID", "SeriesInstanceUID", "DateService", "TimeService"],  # Modality often missing
            "characteristics": ["subtlety"],  # Only subtlety commonly present
            "roi": ["imageSOP_UID", "xCoord", "yCoord"],
            "nodule": ["noduleID"]
        }
    }
    
    # Default structure for unknown cases
    default_expected = {
        "header": ["StudyInstanceUID", "SeriesInstanceUID"],
        "characteristics": [],
        "roi": ["imageSOP_UID"],
        "nodule": ["noduleID"]
    }
    
    return expected_attrs.get(parse_case, default_expected)

def parse_radiology_sample(file_path):
    """
    parse a single radiology xml file and extract nodule/roi data
    
    args:
        file_path: path to the xml file to parse
    
    returns:
        tuple: (main_dataframe, unblinded_dataframe) containing extracted data
    """
    print(f"🔍 Parsing XML file: {os.path.basename(file_path)}")
    
    # detect the parse case first to understand xml structure
    print(f"  📋 Detecting parse case...")
    parse_case = detect_parse_case(file_path)
    print(f"  ✅ Parse case: {parse_case}")
    
    expected_attrs = get_expected_attributes_for_case(parse_case)
    
    # re already imported at module level
    file_id = os.path.basename(file_path).split('.')[0]
    print(f"  📄 File ID: {file_id}")
    
    print(f"  🔄 Loading XML structure...")
    tree = ET.parse(file_path)
    xml_root = tree.getroot()
    print(f"  ✅ XML loaded, root element: {root.tag.split('}')[-1] if '}' in root.tag else root.tag}")
    
    # Debug logging for N/A diagnosis
    debug_info = []

    # Dynamically get the namespace from the root tag
    m = re.match(r'\{(.*)\}', root.tag)
    ns_uri = m.group(1) if m else ''
    
    # Helper to build tag with or without namespace
    def tag(name):
        return f"{{{ns_uri}}}{name}" if ns_uri else name

    # Detect XML structure based on root element
    root_tag_name = root.tag.split('}')[-1] if '}' in root.tag else root.tag
    is_lidc_format = root_tag_name == 'LidcReadMessage'
    
    # Extract header information with expected vs missing logic
    print(f"  🔍 Extracting header information...")
    header = root.find(tag('ResponseHeader'))
    header_values = {}
    
    if header is not None:
        print(f"  ✅ ResponseHeader found")
        debug_info.append("✓ ResponseHeader found")
        
        # Check each expected header field
        for field in ["StudyInstanceUID", "SeriesInstanceUID", "SeriesInstanceUid", "Modality", "DateService", "TimeService"]:
            if field == "SeriesInstanceUID":
                # Handle different spelling variations
                elem = header.find(tag('SeriesInstanceUID')) or header.find(tag('SeriesInstanceUid'))
                field_key = "SeriesInstanceUID"
            else:
                elem = header.find(tag(field))
                field_key = field
            
            if elem is not None and elem.text:
                header_values[field_key] = elem.text
            elif field_key in expected_attrs["header"]:
                header_values[field_key] = "MISSING"
                debug_info.append(f"⚠️  {field_key} expected but MISSING")
            else:
                header_values[field_key] = "#N/A"
    else:
        print(f"  ⚠️  ResponseHeader NOT FOUND")
        debug_info.append("❌ ResponseHeader NOT FOUND")
        # Set all header fields based on expectations
        for field in ["StudyInstanceUID", "SeriesInstanceUID", "Modality", "DateService", "TimeService"]:
            if field in expected_attrs["header"]:
                header_values[field] = "MISSING"
            else:
                header_values[field] = "#N/A"
    
    # Extract values with defaults
    study_uid = header_values.get("StudyInstanceUID", "#N/A")
    series_uid = header_values.get("SeriesInstanceUID", "#N/A") 
    modality = header_values.get("Modality", "#N/A")
    date_service = header_values.get("DateService", "#N/A")
    time_service = header_values.get("TimeService", "#N/A")
    print(f"  📊 Header extracted: StudyUID={study_uid[:20]}...{'(truncated)' if len(study_uid) > 20 else ''}")

    data_rows = []
    unblinded_data_rows = []

    # Determine session element name based on format
    session_tag = 'readingSession' if is_lidc_format else 'CXRreadingSession'
    unblinded_tag = 'unblindedReadNodule' if is_lidc_format else 'unblindedRead'

    # Look for session elements
    print(f"  🔍 Looking for reading sessions...")
    sessions = root.findall(tag(session_tag))
    print(f"  📊 Found {len(sessions)} sessions (searching for {session_tag})")
    debug_info.append(f"Sessions found: {len(sessions)} (looking for {session_tag})")
    
    if not sessions:
        print(f"  ⚠️  No sessions found - trying alternative session tags")
        debug_info.append(f"❌ NO SESSIONS FOUND - trying alternative session tags")
        # Try alternative session tags
        alt_sessions = root.findall(tag('readingSession')) + root.findall(tag('CXRreadingSession'))
        print(f"  📊 Alternative sessions found: {len(alt_sessions)}")
        debug_info.append(f"Alternative sessions: {len(alt_sessions)}")
        if alt_sessions:
            sessions = alt_sessions
            print(f"  ✅ Using alternative sessions")
            debug_info.append("✓ Using alternative sessions")
    
    # Print debug info for files with issues
    if not sessions or any("❌" in info for info in debug_info):
        print(f"\nDEBUG INFO for {file_id}:")
        for info in debug_info:
            print(f"  {info}")
        if not sessions:
            print(f"  Root children: {[child.tag for child in root]}")
    
    print(f"  🔄 Processing {len(sessions)} sessions...")
    for session_idx, session in enumerate(sessions):
        print(f"    📋 Session {session_idx + 1}/{len(sessions)}")
        
        rad_base_elem = session.find(tag('servicingRadiologistID'))
        rad_base = rad_base_elem.text if rad_base_elem is not None else "unknown"
        
        # Use session index + 1 for consistent radiologist numbering
        radiologist = f"anonRad{session_idx + 1}"
        print(f"      👨‍⚕️ Radiologist: {radiologist} (base: {rad_base})")
        
        # Check if this is the last radiologist (unblinded read)
        is_last_radiologist = session_idx == len(sessions) - 1

        # Look for unblinded read elements
        unblinded_reads = session.findall(tag(unblinded_tag))
        print(f"      📊 Found {len(unblinded_reads)} unblinded reads")
        
        for unblinded_idx, unblinded in enumerate(unblinded_reads):
            print(f"        🔍 Processing unblinded read {unblinded_idx + 1}/{len(unblinded_reads)}")
            
            nodule_id_elem = unblinded.find(tag('noduleID'))
            nodule_id = nodule_id_elem.text if nodule_id_elem is not None else "#N/A"
            print(f"          📌 Nodule ID: {nodule_id}")
            
            # Parse characteristics with expected vs missing logic
            print(f"          🔍 Extracting characteristics...")
            characteristics = unblinded.find(tag('characteristics'))
            char_values = {}
            
            if characteristics is not None:
                print(f"          ✅ Characteristics found")
                # Check each characteristic field
                for char_field in ["confidence", "subtlety", "obscuration", "reason"]:
                    elem = characteristics.find(tag(char_field))
                    if elem is not None and elem.text:
                        char_values[char_field] = elem.text
                    elif char_field in expected_attrs["characteristics"]:
                        char_values[char_field] = "MISSING"
                    else:
                        char_values[char_field] = "#N/A"
            else:
                print(f"          ⚠️  No characteristics found")
                # No characteristics found
                for char_field in ["confidence", "subtlety", "obscuration", "reason"]:
                    if char_field in expected_attrs["characteristics"]:
                        char_values[char_field] = "MISSING"
                    else:
                        char_values[char_field] = "#N/A"
            
            confidence = char_values.get("confidence", "#N/A")
            subtlety = char_values.get("subtlety", "#N/A")
            obscuration = char_values.get("obscuration", "#N/A")
            reason = char_values.get("reason", "#N/A")
            print(f"          📊 Extracted: confidence={confidence}, subtlety={subtlety}, obscuration={obscuration}, reason={reason}")

            # Process ROI elements with expected vs missing logic
            print(f"          🔍 Processing ROI elements...")
            rois = unblinded.findall(tag('roi'))
            print(f"          📊 Found {len(rois)} ROI elements")
            
            if not rois:
                print(f"          ⚠️  No ROIs found - creating entry with missing ROI data")
                # No ROIs found - determine what should be marked as MISSING vs N/A
                sop_uid = "MISSING" if "imageSOP_UID" in expected_attrs["roi"] else "#N/A"
                x = "MISSING" if "xCoord" in expected_attrs["roi"] else "#N/A"
                y = "MISSING" if "yCoord" in expected_attrs["roi"] else "#N/A"
                
                # Create one entry for missing ROI data
                row_data = {
                    "FileID": file_id,
                    "ParseCase": parse_case,
                    "Radiologist": radiologist,
                    "SessionType": "Standard",  # No ROIs means standard session
                    "NoduleID": int(nodule_id) if nodule_id not in ["#N/A", "MISSING"] and str(nodule_id).isdigit() else nodule_id,
                    "Confidence": float(confidence) if confidence not in ["#N/A", "MISSING"] and str(confidence).replace('.', '', 1).isdigit() else confidence,
                    "Subtlety": float(subtlety) if subtlety not in ["#N/A", "MISSING"] and str(subtlety).replace('.', '', 1).isdigit() else subtlety,
                    "Obscuration": float(obscuration) if obscuration not in ["#N/A", "MISSING"] and str(obscuration).replace('.', '', 1).isdigit() else obscuration,
                    "Reason": reason,
                    "SOP_UID": sop_uid,
                    "X_coord": x,
                    "Y_coord": y,
                    "CoordCount": 0,  # No coordinates
                    "StudyInstanceUID": study_uid,
                    "SeriesInstanceUID": series_uid,
                    "Modality": modality,
                    "DateService": date_service,
                    "TimeService": time_service
                }
                
                if is_last_radiologist:
                    unblinded_data_rows.append(row_data)
                    print(f"          ✅ Added row to unblinded data")
                else:
                    data_rows.append(row_data)
                    print(f"          ✅ Added row to main data")
            else:
                for roi_idx, roi in enumerate(rois):
                    print(f"            🔍 Processing ROI {roi_idx + 1}/{len(rois)}")
                    # Parse ROI data with expected vs missing logic
                    sop_uid_elem = roi.find(tag('imageSOP_UID'))
                    if sop_uid_elem is not None and sop_uid_elem.text:
                        sop_uid = sop_uid_elem.text
                    elif "imageSOP_UID" in expected_attrs["roi"]:
                        sop_uid = "MISSING"
                    else:
                        sop_uid = "#N/A"
                    
                    # Get coordinates including Z position from edgeMap with expected vs missing logic
                    print(f"            🔍 Extracting coordinates...")
                    x, y, z = "#N/A", "#N/A", "#N/A"  # Default values
                    
                    # First, try to get imageZposition from roi level
                    z_elem = roi.find(tag('imageZposition'))
                    if z_elem is not None and z_elem.text:
                        z = z_elem.text
                        print(f"            📍 Z coordinate from ROI level: {z}")
                    
                    edge_maps = roi.findall(tag('edgeMap'))
                    print(f"            📊 Found {len(edge_maps)} edge maps")
                    
                    if edge_maps:
                        # Use the first edge map for coordinates
                        first_edge = edge_maps[0]
                        x_elem = first_edge.find(tag('xCoord'))
                        y_elem = first_edge.find(tag('yCoord'))
                        
                        # Also try to get z from edgeMap if not found at roi level
                        if z == "#N/A":
                            z_edge_elem = first_edge.find(tag('imageZposition'))
                            if z_edge_elem is not None and z_edge_elem.text:
                                z = z_edge_elem.text
                                print(f"            📍 Z coordinate from edge map: {z}")
                        
                        if x_elem is not None and x_elem.text:
                            x = x_elem.text
                        elif "xCoord" in expected_attrs["roi"]:
                            x = "MISSING"
                            
                        if y_elem is not None and y_elem.text:
                            y = y_elem.text
                        elif "yCoord" in expected_attrs["roi"]:
                            y = "MISSING"
                        
                        print(f"            📍 Coordinates extracted: X={x}, Y={y}, Z={z}")
                    else:
                        # Look for single edge map (original format)
                        edge = roi.find(tag('edgeMap'))
                        if edge is not None:
                            x_elem = edge.find(tag('xCoord'))
                            y_elem = edge.find(tag('yCoord'))
                            
                            # Also try to get z from edgeMap if not found at roi level
                            if z == "#N/A":
                                z_edge_elem = edge.find(tag('imageZposition'))
                                if z_edge_elem is not None and z_edge_elem.text:
                                    z = z_edge_elem.text
                            
                            if x_elem is not None and x_elem.text:
                                x = x_elem.text
                            elif "xCoord" in expected_attrs["roi"]:
                                x = "MISSING"
                                
                            if y_elem is not None and y_elem.text:
                                y = y_elem.text
                            elif "yCoord" in expected_attrs["roi"]:
                                y = "MISSING"
                        else:
                            # No edge map found
                            if "xCoord" in expected_attrs["roi"]:
                                x = "MISSING"
                            if "yCoord" in expected_attrs["roi"]:
                                y = "MISSING"

                    # Count coordinates for this ROI to determine if it's a detailed session
                    edge_maps = roi.findall(tag('edgeMap'))
                    coord_count = len(edge_maps)
                    
                    # Mark sessions with many coordinates as "Detailed"
                    session_type = "Detailed" if coord_count > 10 else "Standard"
                    
                    row_data = {
                        "FileID": file_id,
                        "ParseCase": parse_case,
                        "Radiologist": radiologist,
                        "SessionType": session_type,  # New field to identify detailed sessions
                        "NoduleID": int(nodule_id) if nodule_id not in ["#N/A", "MISSING"] and str(nodule_id).isdigit() else nodule_id,
                        "Confidence": float(confidence) if confidence not in ["#N/A", "MISSING"] and str(confidence).replace('.', '', 1).isdigit() else confidence,
                        "Subtlety": float(subtlety) if subtlety not in ["#N/A", "MISSING"] and str(subtlety).replace('.', '', 1).isdigit() else subtlety,
                        "Obscuration": float(obscuration) if obscuration not in ["#N/A", "MISSING"] and str(obscuration).replace('.', '', 1).isdigit() else obscuration,
                        "Reason": reason,
                        "SOP_UID": sop_uid,
                        "X_coord": float(x) if x not in ["#N/A", "MISSING"] and str(x).replace('.', '', 1).isdigit() else x,
                        "Y_coord": float(y) if y not in ["#N/A", "MISSING"] and str(y).replace('.', '', 1).isdigit() else y,
                        "Z_coord": float(z) if z not in ["#N/A", "MISSING"] and str(z).replace('.', '', 1).isdigit() else z,
                        "CoordCount": coord_count,  # Track number of coordinates
                        "StudyInstanceUID": study_uid,
                        "SeriesInstanceUID": series_uid,
                        "Modality": modality,
                        "DateService": date_service,
                        "TimeService": time_service
                    }
                    
                    if is_last_radiologist:
                        unblinded_data_rows.append(row_data)
                        print(f"            ✅ Added ROI row to unblinded data")
                    else:
                        data_rows.append(row_data)
                        print(f"            ✅ Added ROI row to main data")

    print(f"  🏁 Parsing complete for {file_id}")
    print(f"    📊 Main data rows: {len(data_rows)}")
    print(f"    📊 Unblinded data rows: {len(unblinded_data_rows)}")
    return pd.DataFrame(data_rows), pd.DataFrame(unblinded_data_rows)
def parse_multiple(files):
    """
    parse multiple files with memory optimization and batch processing
    
    args:
        files: list of file paths to parse
        
    returns:
        tuple: (case_data_dict, case_unblinded_data_dict) organized by parse case
    """
    case_data = {}
    case_unblinded_data = {}
    
    total_files = len(files)
    print(f"🚀 Starting to parse {total_files} files...")
    print(f"📁 Files to process:")
    for i, f in enumerate(files, 1):
        print(f"   {i}. {os.path.basename(f)}")
    
    # Process files in batches for memory efficiency
    batch_size = 50 if total_files > 100 else total_files
    print(f"📦 Processing in batches of {batch_size}")
    
    for batch_start in range(0, total_files, batch_size):
        batch_end = min(batch_start + batch_size, total_files)
        batch_files = files[batch_start:batch_end]
        
        print(f"📦 Processing batch {batch_start//batch_size + 1}/{(total_files-1)//batch_size + 1} ({len(batch_files)} files)")
        
        for idx, f in enumerate(batch_files):
            global_idx = batch_start + idx + 1
            try:
                print(f"🔄 Processing file {global_idx}/{total_files}: {os.path.basename(f)}")
                df, unblinded_df = parse_radiology_sample(f)
                
                # Group main data by parse case
                if not df.empty:
                    print(f"  📊 Main data: {len(df)} rows")
                    for case in df['ParseCase'].unique():
                        case_df = df[df['ParseCase'] == case].copy()
                        if case not in case_data:
                            case_data[case] = pd.DataFrame()
                            print(f"  📋 New parse case found: {case}")
                        case_data[case] = pd.concat([case_data[case], case_df], ignore_index=True)
                else:
                    print(f"  ⚠️  Main data is empty")
                
                # Group unblinded data by parse case
                if not unblinded_df.empty:
                    print(f"  📊 Unblinded data: {len(unblinded_df)} rows")
                    for case in unblinded_df['ParseCase'].unique():
                        case_df = unblinded_df[unblinded_df['ParseCase'] == case].copy()
                        if case not in case_unblinded_data:
                            case_unblinded_data[case] = pd.DataFrame()
                        case_unblinded_data[case] = pd.concat([case_unblinded_data[case], case_df], ignore_index=True)
                else:
                    print(f"  ⚠️  Unblinded data is empty")
                        
            except Exception as e:
                print(f"❌ Error parsing {f}: {e}")
        
        # Clear memory after each batch
        gc.collect()
        print(f"✅ Batch {batch_start//batch_size + 1} completed, memory cleared")
    
    print(f"🏁 Completed parsing {total_files} files!")
    
    # Print summary
    print(f"📊 Parsing Summary:")
    print(f"  📋 Main data parse cases: {list(case_data.keys())}")
    print(f"  📋 Unblinded data parse cases: {list(case_unblinded_data.keys())}")
    
    total_main_rows = sum(len(df) for df in case_data.values())
    total_unblinded_rows = sum(len(df) for df in case_unblinded_data.values())
    print(f"  📊 Total main data rows: {total_main_rows}")
    print(f"  📊 Total unblinded data rows: {total_unblinded_rows}")
    print(f"  📊 Grand total: {total_main_rows + total_unblinded_rows} rows")
    
    return case_data, case_unblinded_data


def detect_parse_case(file_path):
    """
    Detect the structure/case of an XML file for appropriate parsing strategy
    """
    try:
        tree = ET.parse(file_path)
        xml_root = tree.getroot()
        # Get namespace if present
        m = re.match(r'\{(.*)\}', xml_root.tag)
        ns_uri = m.group(1) if m else ''
        def tag(name):
            return f"{{{ns_uri}}}{name}" if ns_uri else name
        
        # Check for basic structure indicators
        header = root.find(tag('ResponseHeader'))
        sessions = root.findall(tag('readingSession')) or root.findall(tag('CXRreadingSession'))
        
        if not sessions:
            return "No_Sessions_Found"
        
        # Analyze first session for characteristics
        first_session = sessions[0]
        unblinded_reads = first_session.findall(tag('unblindedReadNodule')) or first_session.findall(tag('unblindedRead'))
        
        if not unblinded_reads:
            return "No_Reads_Found"
        
        # Analyze first read for characteristics
        first_read = unblinded_reads[0]
        characteristics = first_read.find(tag('characteristics'))
        
        if characteristics is None:
            return "No_Characteristics"
        
        # Count available characteristics
        char_fields = ['confidence', 'subtlety', 'obscuration', 'reason']
        available_chars = []
        for field in char_fields:
            elem = characteristics.find(tag(field))
            if elem is not None and elem.text:
                available_chars.append(field)
        
        # Determine case based on available characteristics and header completeness
        header_complete = header is not None
        modality_present = False
        if header:
            modality_elem = header.find(tag('Modality'))
            modality_present = modality_elem is not None and modality_elem.text
        
        # Classification logic
        if len(available_chars) >= 3 and 'reason' in available_chars and header_complete and modality_present:
            return "Complete_Attributes"
        elif 'reason' in available_chars and len(available_chars) >= 2:
            return "With_Reason_Partial"
        elif len(available_chars) >= 2 and 'confidence' in available_chars and 'subtlety' in available_chars:
            return "Core_Attributes_Only"
        elif len(available_chars) == 1:
            return "Minimal_Attributes"
        elif len(sessions) == 1:
            return "LIDC_Single_Session"
        elif len(sessions) == 2:
            return "LIDC_Multi_Session_2"
        elif len(sessions) == 3:
            return "LIDC_Multi_Session_3"
        elif len(sessions) == 4:
            return "LIDC_Multi_Session_4"
        else:
            return "Unknown_Structure"
            
    except Exception as e:
        print(f"Error detecting parse case for {file_path}: {e}")
        return "XML_Parse_Error"


if __name__ == "__main__":
    # This allows the module to be run directly for testing
    root = tk.Tk()
    # NYTXMLGuiApp is undefined; replace with a placeholder or comment
    # app = NYTXMLGuiApp(root)
    # root.mainloop()
