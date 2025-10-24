"""
Excel Exporter for RA-D-PS - Schema-Agnostic Data Export System

This module handles all Excel formatting and export operations, extracting
629+ lines of formatting logic from gui.py into a reusable, testable module.

Supports:
- RA-D-PS format with radiologist blocks
- Template format with color coding
- Multi-sheet exports
- Auto-versioning and naming
- Cross-platform file opening
"""

from pathlib import Path
from typing import Any, Dict, List, Optional, Union
from datetime import datetime
import re
import pandas as pd

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import FormulaRule
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .base import BaseExporter, ExportError
from ..schemas.canonical import RadiologyCanonicalDocument


class ExcelExporter(BaseExporter):
    """
    Base Excel exporter providing core functionality for all Excel exports.
    """
    
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """Initialize Excel exporter with configuration."""
        super().__init__(config)
        
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install with: pip install openpyxl>=3.0.0"
            )
        
        # Default styling colors
        self.blue_argb = self.config.get('blue_color', 'FF4472C4')
        self.light_blue_argb = self.config.get('light_blue_color', 'FFCCE5FF')
        self.white_argb = 'FFFFFFFF'
    
    def export(self, data: Any, output_path: Path, **kwargs) -> Path:
        """
        Export data to Excel file.
        
        Args:
            data: Data to export (format depends on exporter subclass)
            output_path: Path where output should be written
            **kwargs: Additional exporter-specific options
            
        Returns:
            Path to exported file
        """
        raise NotImplementedError("Use specialized subclass (RADPSExcelFormatter, etc.)")
    
    def validate_data(self, data: Any) -> bool:
        """Validate data format for Excel export."""
        return isinstance(data, (list, dict, pd.DataFrame))
    
    def _apply_row_striping(self, ws, non_spacer_cols: List[int], 
                           data_start_row: int = 2, 
                           white_argb: str = None, blue_argb: str = None):
        """
        Apply alternating row colors to non-spacer columns.
        
        Args:
            ws: Worksheet object
            non_spacer_cols: List of column indices (1-based) that aren't spacers
            data_start_row: First row of data (default 2, after header)
            white_argb: White color code (default FFFFFFFF)
            blue_argb: Light blue color code (default FFCCE5FF)
        """
        white_argb = white_argb or self.white_argb
        blue_argb = blue_argb or self.light_blue_argb
        
        white_fill = PatternFill(start_color=white_argb, end_color=white_argb, fill_type="solid")
        blue_fill = PatternFill(start_color=blue_argb, end_color=blue_argb, fill_type="solid")
        
        max_row = ws.max_row
        for row_idx in range(data_start_row, max_row + 1):
            fill = white_fill if (row_idx - data_start_row) % 2 == 0 else blue_fill
            for col_idx in non_spacer_cols:
                ws.cell(row=row_idx, column=col_idx).fill = fill
    
    def _auto_size_columns(self, ws, cols: List[Optional[str]], 
                          min_width: float = 12, max_width: float = 50):
        """
        Auto-size worksheet columns based on content.
        
        Args:
            ws: Worksheet object
            cols: Column headers (None for spacers)
            min_width: Minimum column width
            max_width: Maximum column width
        """
        for col_idx, header in enumerate(cols, start=1):
            if header is None:  # Skip spacer columns
                continue
            
            max_length = len(str(header)) if header else 0
            column_letter = get_column_letter(col_idx)
            
            # Check data cells for max length
            for cell in ws[column_letter]:
                try:
                    cell_length = len(str(cell.value)) if cell.value else 0
                    max_length = max(max_length, cell_length)
                except:
                    pass
            
            # Set column width with constraints
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _fill_spacer_columns(self, ws, cols: List[Optional[str]], 
                            blue_argb: str = None, width: float = 2.0):
        """
        Fill spacer columns with solid blue color.
        
        Args:
            ws: Worksheet object
            cols: Column headers (None indicates spacer)
            blue_argb: Blue color code
            width: Width of spacer columns
        """
        blue_argb = blue_argb or self.blue_argb
        blue_fill = PatternFill(start_color=blue_argb, end_color=blue_argb, fill_type="solid")
        
        for col_idx, header in enumerate(cols, start=1):
            if header is None:  # Spacer column
                column_letter = get_column_letter(col_idx)
                ws.column_dimensions[column_letter].width = width
                
                # Fill all cells in this column
                for row_idx in range(1, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = blue_fill


class RADPSExcelFormatter(ExcelExporter):
    """
    RA-D-PS format Excel exporter with radiologist blocks and spacers.
    
    RA-D-PS Format Structure:
    - Fixed columns: file #, Study UID, [spacer], NoduleID, [spacer]
    - Radiologist blocks: For each radiologist 1-N:
      [Subtlety, Confidence, Obscuration, Reason, Coordinates, SPACER]
    - Auto-versioning with timestamps
    - Alternating row colors
    - Frozen header row
    """
    
    def export(self, records: List[Dict], output_folder: Path, 
               force_blocks: Optional[int] = None, 
               sheet_name: str = "radiology_data") -> Path:
        """
        Export records to RA-D-PS formatted Excel file.
        
        Args:
            records: List of record dicts with 'radiologists' nested structure
            output_folder: Folder where Excel file will be saved
            force_blocks: Force specific number of radiologist blocks (optional)
            sheet_name: Name for the worksheet
            
        Returns:
            Path to created Excel file
            
        Example record format:
            {
                "file_number": "001",
                "study_uid": "1.2.3.4.5",
                "nodule_id": "N1",
                "radiologists": {
                    "1": {"subtlety": 3, "confidence": 4, ...},
                    "2": {"subtlety": 2, "confidence": 5, ...}
                }
            }
        """
        if not self.validate_data(records):
            raise ExportError(f"Invalid data format. Expected list of dicts, got {type(records)}")
        
        if not records:
            raise ExportError("No records to export")
        
        # Determine maximum radiologist count
        R_max = self._get_R_max(records, force_blocks)
        
        # Build column structure
        cols = self._build_columns(R_max)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel limit
        
        # Write headers
        for col_idx, header in enumerate(cols, start=1):
            if header is not None:
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
        
        # Write data rows
        row_idx = 2
        for rec in records:
            # Fixed columns
            ws.cell(row=row_idx, column=1, value=rec.get("file_number"))
            ws.cell(row=row_idx, column=2, value=rec.get("study_uid"))
            ws.cell(row=row_idx, column=4, value=rec.get("nodule_id"))
            
            # Radiologist blocks
            radiologists = rec.get("radiologists", {})
            if radiologists:
                r_keys = sorted(radiologists.keys(), key=lambda x: int(x) if str(x).isdigit() else str(x))
                R_this = len(r_keys)
            else:
                R_this = int(rec.get("radiologist_count", 0))
                r_keys = [str(i) for i in range(1, R_this + 1)]
            
            col_ptr = 6  # Start of first radiologist block
            for r in range(1, R_max + 1):
                if r <= R_this:
                    r_key = r_keys[r - 1] if r <= len(r_keys) else str(r)
                    rdict = radiologists.get(r_key, {})
                    
                    ws.cell(row=row_idx, column=col_ptr, value=rdict.get("subtlety"))
                    ws.cell(row=row_idx, column=col_ptr + 1, value=rdict.get("confidence"))
                    ws.cell(row=row_idx, column=col_ptr + 2, value=rdict.get("obscuration"))
                    ws.cell(row=row_idx, column=col_ptr + 3, value=rdict.get("reason"))
                    ws.cell(row=row_idx, column=col_ptr + 4, value=rdict.get("coordinates"))
                
                col_ptr += 6  # Move to next radiologist block
            
            row_idx += 1
        
        # Apply formatting
        non_spacer_indices = self._get_non_spacer_indices(cols)
        self._apply_row_striping(ws, non_spacer_indices)
        ws.freeze_panes = "A2"  # Freeze header row
        self._auto_size_columns(ws, cols)
        self._fill_spacer_columns(ws, cols)
        
        # Generate output path with auto-naming
        output_path = self._get_output_path(output_folder)
        wb.save(output_path)
        
        return output_path
    
    def validate_data(self, data: Any) -> bool:
        """Validate RA-D-PS format data."""
        if not isinstance(data, list):
            return False
        
        if not data:
            return True  # Empty list is valid
        
        # Check first record structure
        sample = data[0]
        return isinstance(sample, dict) and (
            "radiologists" in sample or 
            "radiologist_count" in sample or
            any(key.startswith("radiologist_") for key in sample.keys())
        )
    
    def _get_R_max(self, records: List[Dict], force_blocks: Optional[int] = None) -> int:
        """
        Determine maximum number of radiologists across all records.
        
        Args:
            records: List of record dicts
            force_blocks: Override with specific count
            
        Returns:
            Maximum radiologist count
        """
        R_max = force_blocks or 0
        
        for rec in records:
            cand = 0
            
            # Check radiologist_count field
            if isinstance(rec.get("radiologist_count"), int):
                cand = max(cand, rec["radiologist_count"])
            
            # Check radiologists dict
            if isinstance(rec.get("radiologists"), dict):
                cand = max(cand, len(rec["radiologists"]))
            
            # Check numbered radiologist fields
            rad_nums = []
            for key in rec.keys():
                match = re.fullmatch(r"radiologist_(\d+)", str(key))
                if match:
                    rad_nums.append(int(match.group(1)))
            if rad_nums:
                cand = max(cand, max(rad_nums))
            
            R_max = max(R_max, cand)
        
        return R_max if R_max > 0 else 1  # Minimum 1 radiologist
    
    def _build_columns(self, R_max: int) -> List[Optional[str]]:
        """
        Build column structure with spacers.
        
        Args:
            R_max: Number of radiologist blocks
            
        Returns:
            List of column headers (None for spacers)
        """
        cols = ["file #", "Study UID", None, "NoduleID", None]
        
        for r in range(1, R_max + 1):
            cols.extend([
                f"R{r} Subtlety",
                f"R{r} Confidence",
                f"R{r} Obscuration",
                f"R{r} Reason",
                f"R{r} Coordinates",
                None  # Spacer after each radiologist block
            ])
        
        return cols
    
    def _get_non_spacer_indices(self, cols: List[Optional[str]]) -> List[int]:
        """Get 1-based indices of non-spacer columns."""
        return [i for i, col in enumerate(cols, start=1) if col is not None]
    
    def _get_output_path(self, output_folder: Path) -> Path:
        """
        Generate auto-named output path with versioning.
        
        Format: {folder_name}_RA-D-PS_{timestamp}.xlsx
        Auto-versions: _v2, _v3, etc. if file exists
        """
        folder_name = self._sanitize_filename(output_folder.name or "export")
        timestamp = self._get_timestamp()
        base_filename = f"{folder_name}_RA-D-PS_{timestamp}.xlsx"
        base_path = output_folder / base_filename
        
        return self._next_versioned_path(base_path)


class TemplateExcelFormatter(ExcelExporter):
    """
    Template format Excel exporter with repeating Radiologist 1-4 columns.
    
    Features:
    - Radiologist columns 1-4 (repeating pattern)
    - Color-coded by radiologist
    - Compact rating strings
    - Alternating row colors
    """
    
    # Radiologist colors (light hues)
    RAD_COLORS = {
        1: "E3F2FD",  # Light Blue
        2: "E8F5E9",  # Light Green
        3: "FFF3E0",  # Light Orange
        4: "F3E5F5"   # Light Purple
    }
    
    def export(self, template_data: List[Dict], output_path: Path) -> Path:
        """
        Export data in template format.
        
        Args:
            template_data: List of dicts with template structure
            output_path: Full path including filename
            
        Returns:
            Path to created file
        """
        if not self.validate_data(template_data):
            raise ExportError(f"Invalid template data format")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Radiology Analysis"
        
        # Define headers
        headers = [
            'FileID', 'NoduleID', 'ParseCase', 'SessionType',
            'Radiologist 1', 'Radiologist 2', 'Radiologist 3', 'Radiologist 4',
            'SOP_UID', 'StudyInstanceUID', 'SeriesInstanceUID',
            'X_coord', 'Y_coord', 'Z_coord', 'CoordCount',
            'Modality', 'DateService', 'TimeService'
        ]
        
        # Write headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.blue_argb, 
                                   end_color=self.blue_argb, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data with color coding
        for row_num, data_row in enumerate(template_data, start=2):
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col)
                cell.value = data_row.get(header, "")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Apply radiologist colors
                if header.startswith('Radiologist '):
                    rad_num = int(header.split()[-1])
                    if col % 2 == 0 and data_row.get(header, "").strip():
                        cell.fill = PatternFill(
                            start_color=self.RAD_COLORS[rad_num],
                            end_color=self.RAD_COLORS[rad_num],
                            fill_type="solid"
                        )
                else:
                    # Alternating light gray for non-radiologist columns
                    if col % 2 == 0:
                        cell.fill = PatternFill(start_color="F8F9FA", 
                                               end_color="F8F9FA", fill_type="solid")
                
                # Highlight MISSING values
                if str(cell.value) == "MISSING":
                    cell.fill = PatternFill(start_color="FFE0B3", 
                                           end_color="FFE0B3", fill_type="solid")
        
        # Auto-fit columns
        self._auto_size_columns(ws, headers)
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
        
        # Save
        output_path = Path(output_path)
        output_path = self._next_versioned_path(output_path)
        wb.save(output_path)
        
        return output_path
    
    def validate_data(self, data: Any) -> bool:
        """Validate template format data."""
        return isinstance(data, list) and (
            not data or isinstance(data[0], dict)
        )


__all__ = ['ExcelExporter', 'RADPSExcelFormatter', 'TemplateExcelFormatter']
