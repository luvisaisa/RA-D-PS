"""
Main entry point for the RA-D-PS XML parsing application.

This module provides a GUI interface for parsing and processing radiology XML files.
"""
import sys
import tkinter as tk
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

# Add src to path for development/testing
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

try:
    from ra_d_ps.gui import NYTXMLGuiApp
except ImportError:
    # Fallback to old import for backward compatibility
    try:
        from XMLPARSE import NYTXMLGuiApp
    except ImportError as e:
        print(f"Import error: {e}")
        print("Make sure the ra_d_ps package is installed or XMLPARSE.py is available.")
        sys.exit(1)


def highlight_na_reason(excel_path):
    """
    Highlight cells with '#N/A' values in the 'Reason' column with red background.
    
    Args:
        excel_path (str): Path to the Excel file to process.
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    # find the column index for "Reason"
    reason_col = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == "Reason":
            reason_col = idx
            break
    if reason_col:
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        for row in ws.iter_rows(min_row=2, min_col=reason_col, max_col=reason_col):
            for cell in row:
                if cell.value == "#N/A":
                    cell.fill = red_fill
    wb.save(excel_path)


def main() -> None:
    """
    Main function to initialize and run the GUI application.
    
    Creates a centered 500x500 window and launches the XML GUI application.
    """
    root = tk.Tk()
    # set window size to 500x500 and center it on the screen
    window_width = 500
    window_height = 500
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x = (screen_width // 2) - (window_width // 2)
    y = (screen_height // 2) - (window_height // 2)
    root.geometry(f"{window_width}x{window_height}+{x}+{y}")
    NYTXMLGuiApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
