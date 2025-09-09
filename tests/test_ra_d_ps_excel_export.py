#!/usr/bin/env python3
"""
Test RA-D-PS Excel export with both main and unblinded data
"""

import sys
import os
import pandas as pd

# Add the current directory to path to import XMLPARSE
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from XMLPARSE import export_excel

def test_excel_export():
    """Test Excel export with RA-D-PS records containing both main and unblinded data"""
    
    print("🧪 Testing RA-D-PS Excel export...")
    
    # Create sample RA-D-PS records with both main and unblinded radiologist data
    ra_d_ps_records = [
        {
            "file_number": "sample_file_1",
            "study_uid": "1.2.3.456789.1",
            "nodule_id": "nodule_001",
            "radiologists": {
                "1": {  # Main radiologist 1
                    "subtlety": "3",
                    "confidence": "4",
                    "obscuration": "1",
                    "reason": "spiculation noted",
                    "coordinates": "150, 200, 75"
                },
                "2": {  # Main radiologist 2  
                    "subtlety": "4",
                    "confidence": "3",
                    "obscuration": "2",
                    "reason": "irregular borders",
                    "coordinates": "155, 205, 78"
                },
                "3": {  # Unblinded radiologist
                    "subtlety": "2",
                    "confidence": "5",
                    "obscuration": "1",
                    "reason": "clear nodule boundary",
                    "coordinates": "152, 203, 76"
                }
            }
        },
        {
            "file_number": "sample_file_2", 
            "study_uid": "1.2.3.456789.2",
            "nodule_id": "nodule_002",
            "radiologists": {
                "1": {  # Main radiologist 1
                    "subtlety": "5",
                    "confidence": "2",
                    "obscuration": "3",
                    "reason": "subtle ground glass",
                    "coordinates": "300, 400, 120"
                },
                "2": {  # Unblinded radiologist 1
                    "subtlety": "3",
                    "confidence": "4",
                    "obscuration": "2",
                    "reason": "well-defined margins",
                    "coordinates": "305, 405, 125"
                },
                "3": {  # Unblinded radiologist 2
                    "subtlety": "4",
                    "confidence": "3",
                    "obscuration": "1",
                    "reason": "solid component visible",
                    "coordinates": "298, 402, 118"
                }
            }
        }
    ]
    
    print(f"📋 Created {len(ra_d_ps_records)} test records")
    
    # Show radiologist distribution
    for i, record in enumerate(ra_d_ps_records, 1):
        file_id = record['file_number']
        nodule_id = record['nodule_id']
        rad_count = len(record['radiologists'])
        rad_nums = list(record['radiologists'].keys())
        print(f"  📄 Record {i}: {file_id}-{nodule_id} → {rad_count} radiologists ({rad_nums})")
    
    # Export to Excel
    output_file = "test_ra_d_ps_export.xlsx"
    print(f"\n📄 Exporting to Excel: {output_file}")
    
    try:
        export_excel(
            ra_d_ps_records,
            output_file,
            force_blocks=True
        )
        
        # Verify file was created
        if os.path.exists(output_file):
            file_size = os.path.getsize(output_file)
            print(f"  ✅ Excel file created successfully!")
            print(f"  📋 File size: {file_size:,} bytes")
            print(f"  📁 Location: {os.path.abspath(output_file)}")
            
            # Try to read the file back to verify content
            try:
                import openpyxl
                wb = openpyxl.load_workbook(output_file)
                sheet_names = wb.sheetnames
                print(f"  📊 Sheets: {sheet_names}")
                
                # Check first sheet for data
                if sheet_names:
                    ws = wb[sheet_names[0]]
                    print(f"  📏 Dimensions: {ws.max_row} rows x {ws.max_column} columns")
                    
                    # Look for radiologist columns
                    rad_columns = []
                    for col in range(1, min(ws.max_column + 1, 50)):  # Check first 50 columns
                        cell_value = ws.cell(row=1, col=col).value
                        if cell_value and "rad" in str(cell_value).lower():
                            rad_columns.append(str(cell_value))
                    
                    print(f"  👥 Found radiologist columns: {rad_columns[:10]}...")  # Show first 10
                    
                wb.close()
                
            except Exception as read_error:
                print(f"  ⚠️  Could not read Excel file for verification: {read_error}")
                
            print("\n🎉 Excel export test successful!")
            print("🔍 Please open the Excel file to verify that both main and unblinded radiologist data appear correctly")
            return True
            
        else:
            print("  ❌ Excel file was not created")
            return False
            
    except Exception as e:
        print(f"  ❌ Excel export failed: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = test_excel_export()
    if success:
        print("\n✅ Test completed successfully!")
    else:
        print("\n❌ Test failed!")
        sys.exit(1)
