## draft 2

# import necessary libraries for data handling, xml parsing, os operations, and gui
import datetime
import gc
import os
import pandas as pd
import platform
import re
import subprocess
import traceback
import xml.etree.ElementTree as ET
from collections import defaultdict
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

# Import the database module
try:
    from radiology_database import RadiologyDatabase
    SQLITE_AVAILABLE = True
except ImportError:
    SQLITE_AVAILABLE = False
    print("⚠️ SQLite database features unavailable - install required packages or check radiology_database.py")

def open_file_cross_platform(file_path):
    """Open a file using the default system application across different platforms"""
    try:
        if platform.system() == "Darwin":  # macOS
            subprocess.run(["open", file_path], check=True)
        elif platform.system() == "Windows":  # Windows
            os.startfile(file_path)
        else:  # Linux and others
            subprocess.run(["xdg-open", file_path], check=True)
    except Exception as e:
        print(f"Could not open file {file_path}: {e}")

# -------- RA-D-PS Excel Exporter --------

def _sanitize_name(name: str) -> str:
    """Keep A-Z a-z 0-9 _ -, replace others with underscore."""
    return re.sub(r"[^A-Za-z0-9_\-]+", "_", name.strip())

def _timestamp() -> str:
    """Return local timestamp YYYY-MM-DD_HHMMSS."""
    return datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")

def _next_versioned_path(folder_path: str, base_filename: str) -> str:
    """
    If base filename exists in folder, append _v2, _v3, ... until unused.
    Returns full path.
    """
    path = os.path.join(folder_path, base_filename)
    if not os.path.exists(path):
        return path
    root, ext = os.path.splitext(base_filename)
    i = 2
    while True:
        candidate = os.path.join(folder_path, f"{root}_v{i}{ext}")
        if not os.path.exists(candidate):
            return candidate
        i += 1

def _count_numbered_keys(d, prefix: str) -> int:
    """Count numbered keys like radiologist_1, radiologist_2, etc."""
    nums = []
    for k in d.keys():
        m = re.fullmatch(rf"{re.escape(prefix)}_(\d+)", str(k))
        if m:
            nums.append(int(m.group(1)))
    return max(nums) if nums else 0

def _get_R_max(records: list, force_blocks=None) -> int:
    """Determine maximum radiologist count across records."""
    R_max = 0
    for rec in records:
        cand = 0
        if isinstance(rec.get("radiologist_count"), int):
            cand = max(cand, rec["radiologist_count"])
        if isinstance(rec.get("radiologists"), dict):
            cand = max(cand, len(rec["radiologists"]))
        cand = max(cand, _count_numbered_keys(rec, "radiologist"))
        R_max = max(R_max, cand)
    if force_blocks is not None:
        R_max = max(R_max, int(force_blocks))
    return R_max

def _build_columns(R_max: int) -> list:
    """
    Build the full column plan as a list where None represents a spacer.
    Fixed columns first, then R_max radiologist blocks each followed by a spacer.
    """
    cols = ["file #", "Study UID", None, "NoduleID", None]
    for r in range(1, R_max + 1):
        cols += [
            f"Radiologist_{r}_Subtlety",
            f"Radiologist_{r}_Confidence",
            f"Radiologist_{r}_Obscuration",
            f"Radiologist_{r}_Reason",
            "Coordinates",
            None,  # spacer between radiologists
        ]
    return cols

def _non_spacer_col_indices(cols: list) -> list:
    """Return a list of 1-based column indices that are NOT spacers (i.e., not None)."""
    return [i + 1 for i, c in enumerate(cols) if c is not None]

def _apply_row_striping(ws, non_spacer_indices: list, blue_argb: str = "FFCCE5FF", white_argb: str = "FFFFFFFF"):
    """Apply alternating row striping via conditional formatting (DifferentialStyle) on non-spacer columns."""
    dxf_blue = DifferentialStyle(fill=PatternFill(start_color=blue_argb, end_color=blue_argb, fill_type="solid"))
    dxf_white = DifferentialStyle(fill=PatternFill(start_color=white_argb, end_color=white_argb, fill_type="solid"))
    rule_blue = Rule(type="expression", dxf=dxf_blue, formula=["MOD(ROW(),2)=0"])
    rule_white = Rule(type="expression", dxf=dxf_white, formula=["MOD(ROW(),2)=1"])
    max_rows = 1048576
    for idx in non_spacer_indices:
        col_letter = get_column_letter(idx)
        rng = f"{col_letter}1:{col_letter}{max_rows}"
        ws.conditional_formatting.add(rng, rule_blue)
        ws.conditional_formatting.add(rng, rule_white)

def _auto_size_columns(ws, cols: list):
    """Auto-size all columns based on max cell length; keep spacers narrow."""
    for i, header in enumerate(cols, start=1):
        col_letter = get_column_letter(i)
        if header is None:
            ws.column_dimensions[col_letter].width = 3  # spacer stays thin
            continue
        max_len = 0
        if header:
            max_len = max(max_len, len(str(header)))
        for cell in ws[col_letter]:
            v = cell.value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = max(10, min(max_len + 2, 60))

def _fill_spacer_columns(ws, cols, blue_argb="FFCCE5FF"):
    """Fill all spacer columns with solid blue background."""
    spacer_fill = PatternFill(start_color=blue_argb, end_color=blue_argb, fill_type="solid")
    for i, header in enumerate(cols, start=1):
        if header is None:
            col_letter = get_column_letter(i)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=i, max_col=i):
                for cell in row:
                    cell.fill = spacer_fill

def _set_column_widths(ws, cols: list):
    """Apply reasonable default widths and widen Reason columns."""
    for i, header in enumerate(cols, start=1):
        if header is None:
            ws.column_dimensions[get_column_letter(i)].width = 3  # narrow spacer
        else:
            if str(header).endswith("_Reason"):
                ws.column_dimensions[get_column_letter(i)].width = 24
            else:
                ws.column_dimensions[get_column_letter(i)].width = 12

def export_excel(records, folder_path, sheet="radiology_data", blue_argb="FFCCE5FF", force_blocks=None):
    """
    Export records to Excel with dynamic radiologist blocks, spacer columns (solid blue fill),
    alternating row striping via conditional formatting, and true auto-sizing.
    Returns output_path.
    
    Args:
        records: list of dicts with required keys:
            - file_number (str/int), study_uid (str), nodule_id (str)
            - EITHER radiologist_count (int) OR radiologists dict "1".."R" -> dict(
                  subtlety, confidence, obscuration, reason, coordinates
              )
        folder_path: where to save the file and derive base name
        sheet: worksheet name (default "radiology_data")
        blue_argb: ARGB color for blue stripe and spacers (default FFCCE5FF)
        force_blocks: minimum number of radiologist blocks to create
    """
    if not isinstance(records, list):
        raise ValueError("records must be a list of dictionaries")

    # Ensure folder exists
    os.makedirs(folder_path, exist_ok=True)

    # Determine R_max for header
    R_max = _get_R_max(records, force_blocks=force_blocks)
    cols = _build_columns(R_max)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet

    # Header row
    for i, header in enumerate(cols, start=1):
        if header is not None:
            ws.cell(row=1, column=i, value=header)

    # Data rows
    r_ptr = 2
    for rec in records:
        ws.cell(row=r_ptr, column=1, value=rec.get("file_number"))
        ws.cell(row=r_ptr, column=2, value=rec.get("study_uid"))
        ws.cell(row=r_ptr, column=4, value=rec.get("nodule_id"))

        if isinstance(rec.get("radiologists"), dict) and rec["radiologists"]:
            try:
                r_keys = sorted(rec["radiologists"].keys(), key=lambda x: int(x))
            except Exception:
                r_keys = sorted(rec["radiologists"].keys())
            R_this = len(r_keys)
        else:
            R_this = int(rec.get("radiologist_count", 0))
            if R_this == 0:
                R_this = _count_numbered_keys(rec, "radiologist")
            r_keys = [str(i) for i in range(1, R_this + 1)]

        c_ptr = 6
        for r in range(1, R_max + 1):
            if r <= R_this:
                if "radiologists" in rec and isinstance(rec["radiologists"], dict):
                    rdict = rec["radiologists"].get(str(r), {})
                else:
                    rdict = rec.get(f"radiologist_{r}", {}) or {}
                ws.cell(row=r_ptr, column=c_ptr + 0, value=rdict.get("subtlety"))
                ws.cell(row=r_ptr, column=c_ptr + 1, value=rdict.get("confidence"))
                ws.cell(row=r_ptr, column=c_ptr + 2, value=rdict.get("obscuration"))
                ws.cell(row=r_ptr, column=c_ptr + 3, value=rdict.get("reason"))
                ws.cell(row=r_ptr, column=c_ptr + 4, value=rdict.get("coordinates"))
            c_ptr += 6

        r_ptr += 1

    # conditional formatting: row striping on non-spacer columns
    non_spacers = _non_spacer_col_indices(cols)
    _apply_row_striping(ws, non_spacers, blue_argb=blue_argb, white_argb="FFFFFFFF")

    # freeze panes
    ws.freeze_panes = "A2"

    # auto-size columns
    _auto_size_columns(ws, cols)

    # fill spacer columns solid blue
    _fill_spacer_columns(ws, cols, blue_argb=blue_argb)

    # Auto-naming
    folder_name = os.path.basename(os.path.abspath(folder_path)) or "export"
    folder_name = _sanitize_name(folder_name)
    ts = _timestamp()
    base_filename = f"{folder_name}_RA-D-PS_{ts}.xlsx"
    out_path = _next_versioned_path(folder_path, base_filename)

    # Save
    wb.save(out_path)
    return out_path

# -------- End RA-D-PS Excel Exporter --------

def convert_parsed_data_to_ra_d_ps_format(dataframes):
    """
    Convert parsed XML DataFrames to RA-D-PS format records.
    
    Args:
        dataframes: tuple of (main_df, unblinded_df) from parse_radiology_sample
                   or dict of {parse_case: df} from parse_multiple
    
    Returns:
        list of records suitable for export_excel function
    """
    records = []
    
    # Handle different input formats
    if isinstance(dataframes, tuple):
        # Single file result: (main_df, unblinded_df)
        main_df, unblinded_df = dataframes
        print(f"🔄 Converting RA-D-PS format...")
        print(f"  📊 Main DataFrame: {len(main_df)} rows")
        print(f"  📊 Unblinded DataFrame: {len(unblinded_df)} rows")
        
        combined_df = pd.concat([main_df, unblinded_df], ignore_index=True) if not main_df.empty or not unblinded_df.empty else pd.DataFrame()
        print(f"  📊 Combined DataFrame: {len(combined_df)} rows")
        
        dfs_to_process = [combined_df] if not combined_df.empty else []
    elif isinstance(dataframes, dict):
        # Multiple files result: {parse_case: df}
        print(f"🔄 Converting multiple DataFrames to RA-D-PS format...")
        dfs_to_process = [df for df in dataframes.values() if not df.empty]
        print(f"  📊 Processing {len(dfs_to_process)} DataFrames")
    elif isinstance(dataframes, pd.DataFrame):
        # Single DataFrame
        print(f"🔄 Converting single DataFrame to RA-D-PS format...")
        dfs_to_process = [dataframes] if not dataframes.empty else []
    else:
        print(f"⚠️  Unknown dataframes format: {type(dataframes)}")
        dfs_to_process = []
    
    for df in dfs_to_process:
        if df.empty:
            continue
            
        print(f"  🔍 Processing DataFrame with {len(df)} rows")
        
        # Group by file and nodule to aggregate radiologist data
        grouped = df.groupby(['FileID', 'NoduleID'])
        print(f"  📋 Found {len(grouped)} unique file/nodule combinations")
        
        for (file_id, nodule_id), group in grouped:
            print(f"    📄 Processing {file_id} - {nodule_id} ({len(group)} rows)")
            
            # Extract study UID from first row
            study_uid = group['StudyInstanceUID'].iloc[0] if 'StudyInstanceUID' in group.columns else "N/A"
            
            # Build radiologists dictionary
            print(f"      👥 Building radiologists dictionary...")
            radiologists = {}
            for idx, row in group.iterrows():
                radiologist = row.get('Radiologist', f'rad_{idx+1}')
                print(f"        👨‍⚕️ Processing radiologist: {radiologist}")
                
                # Extract radiologist number from name (e.g., "anonRad1" -> "1")
                rad_num_match = re.search(r'(\d+)', str(radiologist))
                rad_num = rad_num_match.group(1) if rad_num_match else str(len(radiologists) + 1)
                print(f"        🔢 Extracted rad_num: {rad_num}")
                
                # Build coordinates string
                x_coord = row.get('X_coord', '')
                y_coord = row.get('Y_coord', '')
                z_coord = row.get('Z_coord', '')
                coordinates = f"{x_coord}, {y_coord}, {z_coord}" if any([x_coord, y_coord, z_coord]) else ""
                print(f"        📍 Coordinates: {coordinates}")
                
                radiologists[rad_num] = {
                    "subtlety": row.get('Subtlety', ''),
                    "confidence": row.get('Confidence', ''),
                    "obscuration": row.get('Obscuration', ''),
                    "reason": row.get('Reason', ''),
                    "coordinates": coordinates.strip(", ")
                }
                print(f"        ✅ Added rad_num {rad_num} to dictionary")
            
            print(f"      📝 Final radiologists dictionary has {len(radiologists)} entries: {list(radiologists.keys())}")
            
            record = {
                "file_number": file_id,
                "study_uid": study_uid,
                "nodule_id": nodule_id,
                "radiologists": radiologists
            }
            
            print(f"      📋 Created record with {len(radiologists)} radiologists")
            records.append(record)
    
    return records

# function to parse a single radiology xml file
def get_expected_attributes_for_case(parse_case):
    """define what attributes should be expected for each parse case"""
    expected_attrs = {
        "Complete_Attributes": {
            "header": ["StudyInstanceUID", "SeriesInstanceUID", "Modality", "DateService", "TimeService"],
            "characteristics": ["confidence", "subtlety", "obscuration", "reason"],
            "roi": ["imageSOP_UID", "xCoord", "yCoord"],
            "nodule": ["noduleID"]
        },
        "With_Reason_Partial": {
            "header": ["StudyInstanceUID", "SeriesInstanceUID", "Modality", "DateService", "TimeService"],
            "characteristics": ["reason", "confidence", "subtlety"],  # Partial but has reason
            "roi": ["imageSOP_UID", "xCoord", "yCoord"],
            "nodule": ["noduleID"]
        },
        "Core_Attributes_Only": {
            "header": ["StudyInstanceUID", "SeriesInstanceUID", "Modality"],
            "characteristics": ["confidence", "subtlety", "obscuration"],  # No reason
            "roi": ["imageSOP_UID", "xCoord", "yCoord"],
            "nodule": ["noduleID"]
        },
        "Minimal_Attributes": {
            "header": ["StudyInstanceUID", "SeriesInstanceUID"],
            "characteristics": ["confidence"],  # Only one attribute
            "roi": ["imageSOP_UID"],
            "nodule": ["noduleID"]
        },
        "No_Characteristics": {
            "header": ["StudyInstanceUID", "SeriesInstanceUID"],
            "characteristics": [],  # No characteristics expected
            "roi": ["imageSOP_UID"],
            "nodule": ["noduleID"]
        },
        "LIDC_Single_Session": {
            "header": ["StudyInstanceUID", "SeriesInstanceUID", "DateService", "TimeService"],  # Modality often missing
            "characteristics": ["subtlety"],  # Only subtlety commonly present
            "roi": ["imageSOP_UID", "xCoord", "yCoord"],
            "nodule": ["noduleID"]
        },
        "LIDC_Multi_Session_2": {
            "header": ["StudyInstanceUID", "SeriesInstanceUID", "DateService", "TimeService"],  # Modality often missing
            "characteristics": ["subtlety"],  # Only subtlety commonly present
            "roi": ["imageSOP_UID", "xCoord", "yCoord"],
            "nodule": ["noduleID"]
        },
        "LIDC_Multi_Session_3": {
            "header": ["StudyInstanceUID", "SeriesInstanceUID", "DateService", "TimeService"],  # Modality often missing
            "characteristics": ["subtlety"],  # Only subtlety commonly present
            "roi": ["imageSOP_UID", "xCoord", "yCoord"],
            "nodule": ["noduleID"]
        },
        "LIDC_Multi_Session_4": {
            "header": ["StudyInstanceUID", "SeriesInstanceUID", "DateService", "TimeService"],  # Modality often missing
            "characteristics": ["subtlety"],  # Only subtlety commonly present
            "roi": ["imageSOP_UID", "xCoord", "yCoord"],
            "nodule": ["noduleID"]
        }
    }
    
    # Default structure for unknown cases
    default_expected = {
        "header": ["StudyInstanceUID", "SeriesInstanceUID"],
        "characteristics": [],
        "roi": ["imageSOP_UID"],
        "nodule": ["noduleID"]
    }
    
    return expected_attrs.get(parse_case, default_expected)

def parse_radiology_sample(file_path):
    """
    parse a single radiology xml file and extract nodule/roi data
    
    args:
        file_path: path to the xml file to parse
    
    returns:
        tuple: (main_dataframe, unblinded_dataframe) containing extracted data
    """
    print(f"🔍 Parsing XML file: {os.path.basename(file_path)}")
    
    # detect the parse case first to understand xml structure
    print(f"  📋 Detecting parse case...")
    parse_case = detect_parse_case(file_path)
    print(f"  ✅ Parse case: {parse_case}")
    
    expected_attrs = get_expected_attributes_for_case(parse_case)
    
    import re
    file_id = os.path.basename(file_path).split('.')[0]
    print(f"  📄 File ID: {file_id}")
    
    print(f"  🔄 Loading XML structure...")
    tree = ET.parse(file_path)
    root = tree.getroot()
    print(f"  ✅ XML loaded, root element: {root.tag.split('}')[-1] if '}' in root.tag else root.tag}")
    
    # Debug logging for N/A diagnosis
    debug_info = []

    # Dynamically get the namespace from the root tag
    m = re.match(r'\{(.*)\}', root.tag)
    ns_uri = m.group(1) if m else ''
    
    # Helper to build tag with or without namespace
    def tag(name):
        return f"{{{ns_uri}}}{name}" if ns_uri else name

    # Detect XML structure based on root element
    root_tag_name = root.tag.split('}')[-1] if '}' in root.tag else root.tag
    is_lidc_format = root_tag_name == 'LidcReadMessage'
    
    # Extract header information with expected vs missing logic
    print(f"  🔍 Extracting header information...")
    header = root.find(tag('ResponseHeader'))
    header_values = {}
    
    if header is not None:
        print(f"  ✅ ResponseHeader found")
        debug_info.append("✓ ResponseHeader found")
        
        # Check each expected header field
        for field in ["StudyInstanceUID", "SeriesInstanceUID", "SeriesInstanceUid", "Modality", "DateService", "TimeService"]:
            if field == "SeriesInstanceUID":
                # Handle different spelling variations
                elem = header.find(tag('SeriesInstanceUID')) or header.find(tag('SeriesInstanceUid'))
                field_key = "SeriesInstanceUID"
            else:
                elem = header.find(tag(field))
                field_key = field
            
            if elem is not None and elem.text:
                header_values[field_key] = elem.text
            elif field_key in expected_attrs["header"]:
                header_values[field_key] = "MISSING"
                debug_info.append(f"⚠️  {field_key} expected but MISSING")
            else:
                header_values[field_key] = "#N/A"
    else:
        print(f"  ⚠️  ResponseHeader NOT FOUND")
        debug_info.append("❌ ResponseHeader NOT FOUND")
        # Set all header fields based on expectations
        for field in ["StudyInstanceUID", "SeriesInstanceUID", "Modality", "DateService", "TimeService"]:
            if field in expected_attrs["header"]:
                header_values[field] = "MISSING"
            else:
                header_values[field] = "#N/A"
    
    # Extract values with defaults
    study_uid = header_values.get("StudyInstanceUID", "#N/A")
    series_uid = header_values.get("SeriesInstanceUID", "#N/A") 
    modality = header_values.get("Modality", "#N/A")
    date_service = header_values.get("DateService", "#N/A")
    time_service = header_values.get("TimeService", "#N/A")
    print(f"  📊 Header extracted: StudyUID={study_uid[:20]}...{'(truncated)' if len(study_uid) > 20 else ''}")

    data_rows = []
    unblinded_data_rows = []

    # Determine session element name based on format
    session_tag = 'readingSession' if is_lidc_format else 'CXRreadingSession'
    unblinded_tag = 'unblindedReadNodule' if is_lidc_format else 'unblindedRead'

    # Look for session elements
    print(f"  🔍 Looking for reading sessions...")
    sessions = root.findall(tag(session_tag))
    print(f"  📊 Found {len(sessions)} sessions (searching for {session_tag})")
    debug_info.append(f"Sessions found: {len(sessions)} (looking for {session_tag})")
    
    if not sessions:
        print(f"  ⚠️  No sessions found - trying alternative session tags")
        debug_info.append(f"❌ NO SESSIONS FOUND - trying alternative session tags")
        # Try alternative session tags
        alt_sessions = root.findall(tag('readingSession')) + root.findall(tag('CXRreadingSession'))
        print(f"  📊 Alternative sessions found: {len(alt_sessions)}")
        debug_info.append(f"Alternative sessions: {len(alt_sessions)}")
        if alt_sessions:
            sessions = alt_sessions
            print(f"  ✅ Using alternative sessions")
            debug_info.append("✓ Using alternative sessions")
    
    # Print debug info for files with issues
    if not sessions or any("❌" in info for info in debug_info):
        print(f"\nDEBUG INFO for {file_id}:")
        for info in debug_info:
            print(f"  {info}")
        if not sessions:
            print(f"  Root children: {[child.tag for child in root]}")
    
    print(f"  🔄 Processing {len(sessions)} sessions...")
    for session_idx, session in enumerate(sessions):
        print(f"    📋 Session {session_idx + 1}/{len(sessions)}")
        
        rad_base_elem = session.find(tag('servicingRadiologistID'))
        rad_base = rad_base_elem.text if rad_base_elem is not None else "unknown"
        
        # Use session index + 1 for consistent radiologist numbering
        radiologist = f"anonRad{session_idx + 1}"
        print(f"      👨‍⚕️ Radiologist: {radiologist} (base: {rad_base})")
        
        # Check if this is the last radiologist (unblinded read)
        is_last_radiologist = session_idx == len(sessions) - 1

        # Look for unblinded read elements
        unblinded_reads = session.findall(tag(unblinded_tag))
        print(f"      📊 Found {len(unblinded_reads)} unblinded reads")
        
        for unblinded_idx, unblinded in enumerate(unblinded_reads):
            print(f"        🔍 Processing unblinded read {unblinded_idx + 1}/{len(unblinded_reads)}")
            
            nodule_id_elem = unblinded.find(tag('noduleID'))
            nodule_id = nodule_id_elem.text if nodule_id_elem is not None else "#N/A"
            print(f"          📌 Nodule ID: {nodule_id}")
            
            # Parse characteristics with expected vs missing logic
            print(f"          🔍 Extracting characteristics...")
            characteristics = unblinded.find(tag('characteristics'))
            char_values = {}
            
            if characteristics is not None:
                print(f"          ✅ Characteristics found")
                # Check each characteristic field
                for char_field in ["confidence", "subtlety", "obscuration", "reason"]:
                    elem = characteristics.find(tag(char_field))
                    if elem is not None and elem.text:
                        char_values[char_field] = elem.text
                    elif char_field in expected_attrs["characteristics"]:
                        char_values[char_field] = "MISSING"
                    else:
                        char_values[char_field] = "#N/A"
            else:
                print(f"          ⚠️  No characteristics found")
                # No characteristics found
                for char_field in ["confidence", "subtlety", "obscuration", "reason"]:
                    if char_field in expected_attrs["characteristics"]:
                        char_values[char_field] = "MISSING"
                    else:
                        char_values[char_field] = "#N/A"
            
            confidence = char_values.get("confidence", "#N/A")
            subtlety = char_values.get("subtlety", "#N/A")
            obscuration = char_values.get("obscuration", "#N/A")
            reason = char_values.get("reason", "#N/A")
            print(f"          📊 Extracted: confidence={confidence}, subtlety={subtlety}, obscuration={obscuration}, reason={reason}")

            # Process ROI elements with expected vs missing logic
            print(f"          🔍 Processing ROI elements...")
            rois = unblinded.findall(tag('roi'))
            print(f"          📊 Found {len(rois)} ROI elements")
            
            if not rois:
                print(f"          ⚠️  No ROIs found - creating entry with missing ROI data")
                # No ROIs found - determine what should be marked as MISSING vs N/A
                sop_uid = "MISSING" if "imageSOP_UID" in expected_attrs["roi"] else "#N/A"
                x = "MISSING" if "xCoord" in expected_attrs["roi"] else "#N/A"
                y = "MISSING" if "yCoord" in expected_attrs["roi"] else "#N/A"
                
                # Create one entry for missing ROI data
                row_data = {
                    "FileID": file_id,
                    "ParseCase": parse_case,
                    "Radiologist": radiologist,
                    "SessionType": "Standard",  # No ROIs means standard session
                    "NoduleID": int(nodule_id) if nodule_id not in ["#N/A", "MISSING"] and str(nodule_id).isdigit() else nodule_id,
                    "Confidence": float(confidence) if confidence not in ["#N/A", "MISSING"] and str(confidence).replace('.', '', 1).isdigit() else confidence,
                    "Subtlety": float(subtlety) if subtlety not in ["#N/A", "MISSING"] and str(subtlety).replace('.', '', 1).isdigit() else subtlety,
                    "Obscuration": float(obscuration) if obscuration not in ["#N/A", "MISSING"] and str(obscuration).replace('.', '', 1).isdigit() else obscuration,
                    "Reason": reason,
                    "SOP_UID": sop_uid,
                    "X_coord": x,
                    "Y_coord": y,
                    "CoordCount": 0,  # No coordinates
                    "StudyInstanceUID": study_uid,
                    "SeriesInstanceUID": series_uid,
                    "Modality": modality,
                    "DateService": date_service,
                    "TimeService": time_service
                }
                
                if is_last_radiologist:
                    unblinded_data_rows.append(row_data)
                    print(f"          ✅ Added row to unblinded data")
                else:
                    data_rows.append(row_data)
                    print(f"          ✅ Added row to main data")
            else:
                for roi_idx, roi in enumerate(rois):
                    print(f"            🔍 Processing ROI {roi_idx + 1}/{len(rois)}")
                    # Parse ROI data with expected vs missing logic
                    sop_uid_elem = roi.find(tag('imageSOP_UID'))
                    if sop_uid_elem is not None and sop_uid_elem.text:
                        sop_uid = sop_uid_elem.text
                    elif "imageSOP_UID" in expected_attrs["roi"]:
                        sop_uid = "MISSING"
                    else:
                        sop_uid = "#N/A"
                    
                    # Get coordinates including Z position from edgeMap with expected vs missing logic
                    print(f"            🔍 Extracting coordinates...")
                    x, y, z = "#N/A", "#N/A", "#N/A"  # Default values
                    
                    # First, try to get imageZposition from roi level
                    z_elem = roi.find(tag('imageZposition'))
                    if z_elem is not None and z_elem.text:
                        z = z_elem.text
                        print(f"            📍 Z coordinate from ROI level: {z}")
                    
                    edge_maps = roi.findall(tag('edgeMap'))
                    print(f"            📊 Found {len(edge_maps)} edge maps")
                    
                    if edge_maps:
                        # Use the first edge map for coordinates
                        first_edge = edge_maps[0]
                        x_elem = first_edge.find(tag('xCoord'))
                        y_elem = first_edge.find(tag('yCoord'))
                        
                        # Also try to get z from edgeMap if not found at roi level
                        if z == "#N/A":
                            z_edge_elem = first_edge.find(tag('imageZposition'))
                            if z_edge_elem is not None and z_edge_elem.text:
                                z = z_edge_elem.text
                                print(f"            📍 Z coordinate from edge map: {z}")
                        
                        if x_elem is not None and x_elem.text:
                            x = x_elem.text
                        elif "xCoord" in expected_attrs["roi"]:
                            x = "MISSING"
                            
                        if y_elem is not None and y_elem.text:
                            y = y_elem.text
                        elif "yCoord" in expected_attrs["roi"]:
                            y = "MISSING"
                        
                        print(f"            📍 Coordinates extracted: X={x}, Y={y}, Z={z}")
                    else:
                        # Look for single edge map (original format)
                        edge = roi.find(tag('edgeMap'))
                        if edge is not None:
                            x_elem = edge.find(tag('xCoord'))
                            y_elem = edge.find(tag('yCoord'))
                            
                            # Also try to get z from edgeMap if not found at roi level
                            if z == "#N/A":
                                z_edge_elem = edge.find(tag('imageZposition'))
                                if z_edge_elem is not None and z_edge_elem.text:
                                    z = z_edge_elem.text
                            
                            if x_elem is not None and x_elem.text:
                                x = x_elem.text
                            elif "xCoord" in expected_attrs["roi"]:
                                x = "MISSING"
                                
                            if y_elem is not None and y_elem.text:
                                y = y_elem.text
                            elif "yCoord" in expected_attrs["roi"]:
                                y = "MISSING"
                        else:
                            # No edge map found
                            if "xCoord" in expected_attrs["roi"]:
                                x = "MISSING"
                            if "yCoord" in expected_attrs["roi"]:
                                y = "MISSING"

                    # Count coordinates for this ROI to determine if it's a detailed session
                    edge_maps = roi.findall(tag('edgeMap'))
                    coord_count = len(edge_maps)
                    
                    # Mark sessions with many coordinates as "Detailed"
                    session_type = "Detailed" if coord_count > 10 else "Standard"
                    
                    row_data = {
                        "FileID": file_id,
                        "ParseCase": parse_case,
                        "Radiologist": radiologist,
                        "SessionType": session_type,  # New field to identify detailed sessions
                        "NoduleID": int(nodule_id) if nodule_id not in ["#N/A", "MISSING"] and str(nodule_id).isdigit() else nodule_id,
                        "Confidence": float(confidence) if confidence not in ["#N/A", "MISSING"] and str(confidence).replace('.', '', 1).isdigit() else confidence,
                        "Subtlety": float(subtlety) if subtlety not in ["#N/A", "MISSING"] and str(subtlety).replace('.', '', 1).isdigit() else subtlety,
                        "Obscuration": float(obscuration) if obscuration not in ["#N/A", "MISSING"] and str(obscuration).replace('.', '', 1).isdigit() else obscuration,
                        "Reason": reason,
                        "SOP_UID": sop_uid,
                        "X_coord": float(x) if x not in ["#N/A", "MISSING"] and str(x).replace('.', '', 1).isdigit() else x,
                        "Y_coord": float(y) if y not in ["#N/A", "MISSING"] and str(y).replace('.', '', 1).isdigit() else y,
                        "Z_coord": float(z) if z not in ["#N/A", "MISSING"] and str(z).replace('.', '', 1).isdigit() else z,
                        "CoordCount": coord_count,  # Track number of coordinates
                        "StudyInstanceUID": study_uid,
                        "SeriesInstanceUID": series_uid,
                        "Modality": modality,
                        "DateService": date_service,
                        "TimeService": time_service
                    }
                    
                    if is_last_radiologist:
                        unblinded_data_rows.append(row_data)
                        print(f"            ✅ Added ROI row to unblinded data")
                    else:
                        data_rows.append(row_data)
                        print(f"            ✅ Added ROI row to main data")

    print(f"  🏁 Parsing complete for {file_id}")
    print(f"    📊 Main data rows: {len(data_rows)}")
    print(f"    📊 Unblinded data rows: {len(unblinded_data_rows)}")
    return pd.DataFrame(data_rows), pd.DataFrame(unblinded_data_rows)
def parse_multiple(files):
    """
    parse multiple files with memory optimization and batch processing
    
    args:
        files: list of file paths to parse
        
    returns:
        tuple: (case_data_dict, case_unblinded_data_dict) organized by parse case
    """
    case_data = {}
    case_unblinded_data = {}
    
    total_files = len(files)
    print(f"🚀 Starting to parse {total_files} files...")
    print(f"📁 Files to process:")
    for i, f in enumerate(files, 1):
        print(f"   {i}. {os.path.basename(f)}")
    
    # Process files in batches for memory efficiency
    batch_size = 50 if total_files > 100 else total_files
    print(f"📦 Processing in batches of {batch_size}")
    
    for batch_start in range(0, total_files, batch_size):
        batch_end = min(batch_start + batch_size, total_files)
        batch_files = files[batch_start:batch_end]
        
        print(f"📦 Processing batch {batch_start//batch_size + 1}/{(total_files-1)//batch_size + 1} ({len(batch_files)} files)")
        
        for idx, f in enumerate(batch_files):
            global_idx = batch_start + idx + 1
            try:
                print(f"🔄 Processing file {global_idx}/{total_files}: {os.path.basename(f)}")
                df, unblinded_df = parse_radiology_sample(f)
                
                # Group main data by parse case
                if not df.empty:
                    print(f"  📊 Main data: {len(df)} rows")
                    for case in df['ParseCase'].unique():
                        case_df = df[df['ParseCase'] == case].copy()
                        if case not in case_data:
                            case_data[case] = pd.DataFrame()
                            print(f"  📋 New parse case found: {case}")
                        case_data[case] = pd.concat([case_data[case], case_df], ignore_index=True)
                else:
                    print(f"  ⚠️  Main data is empty")
                
                # Group unblinded data by parse case
                if not unblinded_df.empty:
                    print(f"  📊 Unblinded data: {len(unblinded_df)} rows")
                    for case in unblinded_df['ParseCase'].unique():
                        case_df = unblinded_df[unblinded_df['ParseCase'] == case].copy()
                        if case not in case_unblinded_data:
                            case_unblinded_data[case] = pd.DataFrame()
                        case_unblinded_data[case] = pd.concat([case_unblinded_data[case], case_df], ignore_index=True)
                else:
                    print(f"  ⚠️  Unblinded data is empty")
                        
            except Exception as e:
                print(f"❌ Error parsing {f}: {e}")
        
        # Clear memory after each batch
        gc.collect()
        print(f"✅ Batch {batch_start//batch_size + 1} completed, memory cleared")
    
    print(f"🏁 Completed parsing {total_files} files!")
    
    # Print summary
    print(f"📊 Parsing Summary:")
    print(f"  📋 Main data parse cases: {list(case_data.keys())}")
    print(f"  📋 Unblinded data parse cases: {list(case_unblinded_data.keys())}")
    
    total_main_rows = sum(len(df) for df in case_data.values())
    total_unblinded_rows = sum(len(df) for df in case_unblinded_data.values())
    print(f"  📊 Total main data rows: {total_main_rows}")
    print(f"  📊 Total unblinded data rows: {total_unblinded_rows}")
    print(f"  📊 Grand total: {total_main_rows + total_unblinded_rows} rows")
    
    return case_data, case_unblinded_data

class NYTXMLGuiApp:
    """
    main gui application for xml parsing and excel export
    provides file/folder selection, parsing, and formatted excel output
    """
    def __init__(self, master: tk.Tk):
        """
        initialize the gui application
        
        args:
            master: the root tkinter window
        """
        self.master = master
        self.master.title("NYT XML Parser")
        self.master.configure(bg="#d7e3fc")
        
        # Set window size to 500x500 and center it
        window_width = 500
        window_height = 500
        
        # Update the window to get accurate screen dimensions
        self.master.update_idletasks()
        screen_width = self.master.winfo_screenwidth()
        screen_height = self.master.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        
        # Set geometry and ensure window is not resizable for consistent layout
        self.master.geometry(f"{window_width}x{window_height}+{x}+{y}")
        self.master.resizable(False, False)
        
        self._create_widgets()
        self.files = []
        self.excel_path = None
        # schedule the signature popup after the window is drawn
        self.master.after(800, self.show_creator_signature)

    def _create_widgets(self):
        """create all gui widgets and layout with consistent styling"""
        font = ("Aptos", 11, "normal")  # aptos (body) font for all widgets

        frame = tk.Frame(self.master, bg="#d7e3fc")  # main frame with background color
        frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)  # pack frame

        # button to select xml files
        btn_files = tk.Button(frame, text="Select XML Files", command=self.select_files, font=font, bg="#d7e3fc", fg="black")
        btn_files.pack(fill=tk.X)

        # button to select folders containing xml files (unified method)
        btn_folder = tk.Button(frame, text="Select Folders", command=self.select_folders, font=font, bg="#d7e3fc", fg="black")
        btn_folder.pack(fill=tk.X)

        # button to select an existing excel file to append to
        btn_excel = tk.Button(frame, text="Select Excel to Append", command=self.select_excel, font=font, bg="#d7e3fc", fg="black")
        btn_excel.pack(fill=tk.X)

        # Export options frame
        export_frame = tk.Frame(frame, bg="#d7e3fc")
        export_frame.pack(fill=tk.X, pady=5)

        # button to export to RA-D-PS Excel format
        btn_export_new = tk.Button(export_frame, text="Export to Excel", command=self.export_ra_d_ps_excel, font=font, bg="#d7e3fc", fg="black")
        btn_export_new.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 2))

        # button to export to SQLite database (if available)
        if SQLITE_AVAILABLE:
            btn_export_sqlite = tk.Button(export_frame, text="Export to SQLite", command=self.export_to_sqlite, font=font, bg="#4CAF50", fg="black")
            btn_export_sqlite.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(2, 0))

        # listbox to show selected xml files
        self.listbox = tk.Listbox(frame, height=6, font=font, bg="#f4f8ff")
        self.listbox.pack(fill=tk.BOTH, expand=True, pady=5)

        # button to append parsed data to selected excel file
        parse_btn = tk.Button(frame, text="Append to Selected Excel", command=self.parse_files, font=font, bg="#d7e3fc", fg="black")
        parse_btn.pack(fill=tk.X)

        # button to clear the file list
        clear_btn = tk.Button(frame, text="Clear File List", command=self.clear_files, font=font, bg="#d7e3fc", fg="black")
        clear_btn.pack(fill=tk.X)

    def select_files(self) -> None:
        """open file dialog to select individual xml files for processing"""
        filenames = filedialog.askopenfilenames(
            title="Choose XML files", filetypes=[("XML Files", "*.xml")]
        )
        if filenames:
            self.files = sorted(list(filenames))
            self._update_file_list()

    def select_folders(self) -> None:
        """unified folder selection with choice between single and multiple folder processing modes"""
        
        # Create a choice dialog
        choice_window = tk.Toplevel(self.master)
        choice_window.title("Select Processing Mode")
        choice_window.geometry("700x300")  # Wider to accommodate three buttons
        choice_window.configure(bg="#d7e3fc")
        choice_window.transient(self.master)
        choice_window.grab_set()
        
        # center the dialog on the main window
        choice_window.update_idletasks()
        main_x = self.master.winfo_x()
        main_y = self.master.winfo_y()
        main_width = self.master.winfo_width()
        main_height = self.master.winfo_height()
        
        dialog_width = choice_window.winfo_width()
        dialog_height = choice_window.winfo_height()
        
        x = main_x + (main_width - dialog_width) // 2
        y = main_y + (main_height - dialog_height) // 2
        choice_window.geometry(f"+{x}+{y}")
        
        font = ("Aptos", 11, "normal")
        
        tk.Label(choice_window, text="Choose processing mode:", font=font, bg="#d7e3fc").pack(pady=10)
        
        def single_folder_mode():
            choice_window.destroy()
            self.select_folder()
        
        def multiple_folders_one_excel_mode():
            choice_window.destroy()
            self.select_multiple_folders_for_one_excel()
        
        def multiple_folders_separate_excel_mode():
            choice_window.destroy()
            self.select_multiple_folders_for_separate_files()
        
        # Buttons for the three modes
        btn_frame = tk.Frame(choice_window, bg="#d7e3fc")
        btn_frame.pack(pady=20)
        
        tk.Button(btn_frame, text="Single Folder\n(All files in one Excel)", 
                 command=single_folder_mode, font=font, bg="#d7e3fc", fg="black", width=20, height=3).pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="Multiple Folders\n(One Excel, sheets per folder)", 
                 command=multiple_folders_one_excel_mode, font=font, bg="#d7e3fc", fg="black", width=20, height=3).pack(side=tk.LEFT, padx=5)
        
        tk.Button(btn_frame, text="Multiple Folders\n(Separate Excel per folder)", 
                 command=multiple_folders_separate_excel_mode, font=font, bg="#d7e3fc", fg="black", width=20, height=3).pack(side=tk.LEFT, padx=5)
        
        tk.Label(choice_window, text="Single: One folder → One Excel\nMultiple+Sheets: Multiple folders → One Excel with separate sheets\nMultiple+Files: Multiple folders → Separate Excel per folder", 
                font=("Aptos", 9, "normal"), bg="#d7e3fc").pack(pady=10)

    def select_folder(self) -> None:
        """select a single folder and add all xml files from it to the processing list"""
        folder = filedialog.askdirectory(title="Choose folder with XML files")
        if folder:
            try:
                # efficient xml file filtering with validation
                xml_files = []
                for f in os.listdir(folder):
                    if (f.lower().endswith(".xml") and 
                        not f.startswith("._") and  # filter out macos resource forks
                        not f.startswith("~") and   # filter out temporary files
                        os.path.isfile(os.path.join(folder, f))):
                        xml_files.append(os.path.join(folder, f))
                
                if not xml_files:
                    # provide helpful feedback about folder contents
                    all_files = [f for f in os.listdir(folder) 
                               if os.path.isfile(os.path.join(folder, f)) and not f.startswith(".")]
                    if all_files:
                        self.show_temporary_error(f"No XML files found in the selected folder.\nFound {len(all_files)} other files.")
                    else:
                        self.show_temporary_error("The selected folder is empty.")
                    return
                
                # sort alphabetically for consistent ordering (works well with zero-padded filenames)
                self.files = sorted(xml_files)
                self._update_file_list()
                self.show_temporary_error(f"Added {len(xml_files)} XML files from folder.")
                
            except PermissionError:
                self.show_temporary_error("Permission denied: Cannot access the selected folder.")
            except OSError as e:
                self.show_temporary_error(f"System error accessing folder: {str(e)}")
            except Exception as e:
                self.show_temporary_error(f"Unexpected error: {str(e)}")
    
    def select_multiple_folders_for_one_excel(self) -> None:
        """
        Select multiple folders and combine all into one Excel with separate sheets per folder
        """
        # Create a custom dialog for multi-folder selection
        multi_folder_window = tk.Toplevel(self.master)
        multi_folder_window.title("Select Multiple Folders - One Excel with Sheets")
        multi_folder_window.geometry("600x400")
        multi_folder_window.configure(bg="#d7e3fc")
        multi_folder_window.transient(self.master)
        multi_folder_window.grab_set()
        
        # Center the dialog
        multi_folder_window.update_idletasks()
        x = (multi_folder_window.winfo_screenwidth() - multi_folder_window.winfo_width()) // 2
        y = (multi_folder_window.winfo_screenheight() - multi_folder_window.winfo_height()) // 2
        multi_folder_window.geometry(f"+{x}+{y}")
        
        font = ("Aptos", 11, "normal")
        folders = []
        
        # Instructions
        tk.Label(multi_folder_window, text="Select multiple folders to combine into one Excel file with separate sheets", 
                font=font, bg="#d7e3fc").pack(pady=10)
        
        # Listbox to show selected folders
        listbox_frame = tk.Frame(multi_folder_window, bg="#d7e3fc")
        listbox_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        folder_listbox = tk.Listbox(listbox_frame, font=font, bg="#f4f8ff")
        folder_listbox.pack(fill=tk.BOTH, expand=True)
        
        # Button frame
        btn_frame = tk.Frame(multi_folder_window, bg="#d7e3fc")
        btn_frame.pack(pady=10)
        
        def add_folder():
            folder = filedialog.askdirectory(title="Select folder to add")
            if folder and folder not in folders:
                folders.append(folder)
                folder_listbox.insert(tk.END, os.path.basename(folder))
        
        def remove_folder():
            selection = folder_listbox.curselection()
            if selection:
                index = selection[0]
                folders.pop(index)
                folder_listbox.delete(index)
        
        def process_folders():
            if not folders:
                messagebox.showwarning("No Folders", "Please select at least one folder.")
                return
            
            multi_folder_window.destroy()
            self._process_multiple_folders_one_excel(folders)
        
        tk.Button(btn_frame, text="Add Folder", command=add_folder, font=font, bg="#d7e3fc").pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Remove Selected", command=remove_folder, font=font, bg="#d7e3fc").pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Process Folders", command=process_folders, font=font, bg="#d7e3fc").pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Cancel", command=multi_folder_window.destroy, font=font, bg="#d7e3fc").pack(side=tk.LEFT, padx=5)

    def select_multiple_folders_for_separate_files(self) -> None:
        """
        select multiple folders at once and create separate excel files for each
        uses native system dialogs for efficient multi-selection on macos
        """
        # Create a custom dialog for multi-folder selection
        multi_folder_window = tk.Toplevel(self.master)
        multi_folder_window.title("Select Multiple Folders")
        multi_folder_window.geometry("600x400")
        multi_folder_window.configure(bg="#d7e3fc")
        multi_folder_window.transient(self.master)
        multi_folder_window.grab_set()
        
        # Center the dialog
        multi_folder_window.update_idletasks()
        x = (multi_folder_window.winfo_screenwidth() - multi_folder_window.winfo_width()) // 2
        y = (multi_folder_window.winfo_screenheight() - multi_folder_window.winfo_height()) // 2
        multi_folder_window.geometry(f"+{x}+{y}")
        
        font = ("Aptos", 11, "normal")
        folders = []
        
        # Instructions
        tk.Label(multi_folder_window, text="Select multiple folders to process:", 
                font=font, bg="#d7e3fc").pack(pady=10)
        
        # Listbox to show selected folders
        listbox_frame = tk.Frame(multi_folder_window, bg="#d7e3fc")
        listbox_frame.pack(pady=10, padx=20, fill=tk.BOTH, expand=True)
        
        folder_listbox = tk.Listbox(listbox_frame, font=font, bg="#f4f8ff")
        scrollbar = tk.Scrollbar(listbox_frame, orient="vertical")
        folder_listbox.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=folder_listbox.yview)
        
        folder_listbox.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        def add_folder():
            # Use native system dialog to select multiple folders
            try:
                import subprocess
                import platform
                
                if platform.system() == "Darwin":  # macOS
                    # Use AppleScript to open Finder with multi-selection
                    script = '''
                    tell application "Finder"
                        activate
                        set selectedFolders to choose folder with prompt "Select folders to process (hold Cmd to select multiple):" with multiple selections allowed
                        set folderPaths to {}
                        repeat with aFolder in selectedFolders
                            set end of folderPaths to POSIX path of aFolder
                        end repeat
                        return folderPaths
                    end tell
                    '''
                    
                    result = subprocess.run(['osascript', '-e', script], 
                                          capture_output=True, text=True)
                    
                    if result.returncode == 0 and result.stdout.strip():
                        # parse the applescript output safely
                        paths_output = result.stdout.strip()
                        # remove the outer braces and split by comma
                        if paths_output.startswith('{') and paths_output.endswith('}'):
                            paths_output = paths_output[1:-1]
                        
                        selected_paths = []
                        if paths_output and paths_output != "false":  # handle user cancellation
                            # split by comma and clean up each path
                            raw_paths = paths_output.split(', ')
                            for path in raw_paths:
                                cleaned_path = path.strip().strip('"')
                                if cleaned_path and os.path.exists(cleaned_path) and os.path.isdir(cleaned_path):
                                    selected_paths.append(cleaned_path)
                        
                        # add selected folders to the list
                        for folder_path in selected_paths:
                            if folder_path not in folders:
                                folders.append(folder_path)
                                # show readable path with ellipsis for long paths
                                display_name = folder_path
                                if len(display_name) > 60:
                                    display_name = "..." + display_name[-57:]
                                folder_listbox.insert(tk.END, display_name)
                        
                        if selected_paths:
                            messagebox.showinfo("Folders Added", f"Added {len(selected_paths)} folder(s) successfully!")
                    
                else:
                    # For non-macOS systems, fall back to multiple single selections
                    while True:
                        folder = filedialog.askdirectory(title="Select a folder (Cancel when done)")
                        if not folder:
                            break
                        if folder not in folders:
                            folders.append(folder)
                            # Show full path with ellipsis if too long
                            display_name = folder
                            if len(display_name) > 60:
                                display_name = "..." + display_name[-57:]
                            folder_listbox.insert(tk.END, display_name)
                        
                        # Ask if they want to add more
                        if not messagebox.askyesno("Add More?", "Do you want to add another folder?"):
                            break
                            
            except Exception as e:
                # Fallback to standard single folder selection
                messagebox.showwarning("Multi-select failed", f"Multi-select not available: {e}\nUsing single selection instead.")
                folder = filedialog.askdirectory(title="Select a folder to add")
                if folder and folder not in folders:
                    folders.append(folder)
                    # Show full path with ellipsis if too long
                    display_name = folder
                    if len(display_name) > 60:
                        display_name = "..." + display_name[-57:]
                    folder_listbox.insert(tk.END, display_name)
        
        def remove_selected():
            selection = folder_listbox.curselection()
            if selection:
                index = selection[0]
                folders.pop(index)
                folder_listbox.delete(index)
        
        def clear_all():
            folders.clear()
            folder_listbox.delete(0, tk.END)
        
        # Button frame
        button_frame = tk.Frame(multi_folder_window, bg="#d7e3fc")
        button_frame.pack(pady=10)
        
        tk.Button(button_frame, text="Add Folders", command=add_folder, 
                 font=font, bg="#d7e3fc", fg="black", width=12).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Remove Selected", command=remove_selected, 
                 font=font, bg="#d7e3fc", fg="black", width=12).pack(side=tk.LEFT, padx=5)
        tk.Button(button_frame, text="Clear All", command=clear_all, 
                 font=font, bg="#d7e3fc", fg="black", width=12).pack(side=tk.LEFT, padx=5)
        
        # Process button frame
        process_frame = tk.Frame(multi_folder_window, bg="#d7e3fc")
        process_frame.pack(pady=10)
        
        def process_folders():
            if not folders:
                messagebox.showwarning("No Folders", "Please select at least one folder.")
                return
            multi_folder_window.destroy()
            self._process_multiple_folders(folders)
        
        def cancel():
            multi_folder_window.destroy()
        
        tk.Button(process_frame, text="Process Selected Folders", command=process_folders, 
                 font=("Aptos", 11, "bold"), bg="#4CAF50", fg="black", width=20).pack(side=tk.LEFT, padx=10)
        tk.Button(process_frame, text="Cancel", command=cancel, 
                 font=font, bg="#f44336", fg="black", width=10).pack(side=tk.LEFT, padx=10)
        
        # Instructions at bottom
        tk.Label(multi_folder_window, text="Click 'Add Folders' to open Finder and select multiple folders at once (hold Cmd to multi-select).\nEach folder will get its own Excel file.", 
                font=("Aptos", 9, "normal"), bg="#d7e3fc").pack(pady=10)
    
    def _process_multiple_folders_one_excel(self, folders):
        """Process multiple folders and combine all data into one Excel with separate sheets per folder"""
        if not folders:
            return
        
        # Ask where to save the combined Excel file
        output_path = filedialog.asksaveasfilename(
            title="Save Combined Excel File",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if not output_path:
            return
        
        # Progress dialog
        progress_window = tk.Toplevel(self.master)
        progress_window.title("Processing Multiple Folders - One Excel")
        progress_window.geometry("600x400")
        progress_window.configure(bg="#d7e3fc")
        progress_window.transient(self.master)
        progress_window.grab_set()
        
        # Center the dialog
        progress_window.update_idletasks()
        x = (progress_window.winfo_screenwidth() - progress_window.winfo_width()) // 2
        y = (progress_window.winfo_screenheight() - progress_window.winfo_height()) // 2
        progress_window.geometry(f"+{x}+{y}")
        
        # Progress text area
        text_frame = tk.Frame(progress_window, bg="#d7e3fc")
        text_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        progress_text = tk.Text(text_frame, bg="#f8f9fa", font=("Courier", 10))
        scrollbar = tk.Scrollbar(text_frame, command=progress_text.yview)
        progress_text.config(yscrollcommand=scrollbar.set)
        progress_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        def log_message(msg):
            progress_text.insert(tk.END, f"{msg}\n")
            progress_text.see(tk.END)
            progress_window.update()
        
        try:
            from openpyxl import Workbook
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)
            
            total_folders = len(folders)
            total_records = 0
            
            log_message(f"🚀 Starting processing of {total_folders} folders...")
            log_message(f"📄 Output file: {os.path.basename(output_path)}")
            log_message("=" * 60)
            
            for folder_idx, folder in enumerate(folders, 1):
                folder_name = os.path.basename(folder)
                log_message(f"\n📁 Processing folder {folder_idx}/{total_folders}: {folder_name}")
                
                # Get XML files from folder
                xml_files = []
                try:
                    for f in os.listdir(folder):
                        if (f.lower().endswith(".xml") and 
                            not f.startswith("._") and
                            os.path.isfile(os.path.join(folder, f))):
                            xml_files.append(os.path.join(folder, f))
                    
                    if not xml_files:
                        log_message(f"   ⚠️ No XML files found in {folder_name}")
                        continue
                    
                    log_message(f"   📊 Found {len(xml_files)} XML files")
                    
                    # Parse files
                    case_data, case_unblinded_data = parse_multiple(xml_files)
                    
                    if not case_data and not case_unblinded_data:
                        log_message(f"   ❌ No data parsed from {folder_name}")
                        continue
                    
                    # Convert to RA-D-PS format
                    combined_dataframes = {}
                    combined_dataframes.update(case_data)
                    combined_dataframes.update(case_unblinded_data)
                    
                    ra_d_ps_records = convert_parsed_data_to_ra_d_ps_format(combined_dataframes)
                    
                    if not ra_d_ps_records:
                        log_message(f"   ❌ No RA-D-PS records generated from {folder_name}")
                        continue
                    
                    folder_record_count = len(ra_d_ps_records)
                    total_records += folder_record_count
                    log_message(f"   ✅ Generated {folder_record_count} RA-D-PS records")
                    
                    # Create worksheet for this folder
                    sanitized_name = re.sub(r"[^A-Za-z0-9_\-]+", "_", folder_name)[:31]  # Excel sheet name limit
                    ws = wb.create_sheet(title=sanitized_name)
                    
                    # Determine R_max for this folder
                    R_max = _get_R_max(ra_d_ps_records)
                    cols = _build_columns(R_max)
                    
                    # Write headers
                    for i, header in enumerate(cols, start=1):
                        if header is not None:
                            ws.cell(row=1, column=i, value=header)
                    
                    # Write data
                    row_idx = 2
                    for rec in ra_d_ps_records:
                        # Fixed fields
                        ws.cell(row=row_idx, column=1, value=rec.get("file_number"))
                        ws.cell(row=row_idx, column=2, value=rec.get("study_uid"))
                        ws.cell(row=row_idx, column=4, value=rec.get("nodule_id"))
                        
                        # Radiologist data
                        if isinstance(rec.get("radiologists"), dict):
                            r_keys = sorted(rec["radiologists"].keys(), key=lambda x: int(x) if str(x).isdigit() else str(x))
                            R_this = len(r_keys)
                        else:
                            R_this = int(rec.get("radiologist_count", 0))
                            r_keys = [str(i) for i in range(1, R_this + 1)]
                        
                        col_ptr = 6  # Start of first radiologist block
                        for r in range(1, R_max + 1):
                            if r <= R_this:
                                r_key = r_keys[r - 1]
                                rdict = rec.get("radiologists", {}).get(r_key, {}) if "radiologists" in rec else rec.get(f"radiologist_{r}", {})
                                ws.cell(row=row_idx, column=col_ptr, value=rdict.get("subtlety"))
                                ws.cell(row=row_idx, column=col_ptr + 1, value=rdict.get("confidence"))
                                ws.cell(row=row_idx, column=col_ptr + 2, value=rdict.get("obscuration"))
                                ws.cell(row=row_idx, column=col_ptr + 3, value=rdict.get("reason"))
                                ws.cell(row=row_idx, column=col_ptr + 4, value=rdict.get("coordinates"))
                            col_ptr += 6  # Move to next radiologist block
                        
                        row_idx += 1
                    
                    # Apply formatting to this sheet
                    non_spacer_indices = _non_spacer_col_indices(cols)
                    _apply_row_striping(ws, non_spacer_indices, blue_argb="FFCCE5FF", white_argb="FFFFFFFF")
                    ws.freeze_panes = "A2"
                    _set_column_widths(ws, cols)
                    
                    log_message(f"   📋 Created sheet '{sanitized_name}' with {folder_record_count} records")
                    
                except Exception as e:
                    log_message(f"   ❌ Error processing {folder_name}: {str(e)}")
                    continue
            
            if total_records > 0:
                # Save the workbook
                wb.save(output_path)
                log_message("\n" + "=" * 60)
                log_message(f"✅ COMBINED EXCEL EXPORT COMPLETED!")
                log_message(f"📊 Total records across all folders: {total_records}")
                log_message(f"📄 Saved to: {output_path}")
                
                # Success message
                success_msg = (
                    f"✅ Combined Excel export completed!\n\n"
                    f"📊 Summary:\n"
                    f"• Processed {total_folders} folders\n"
                    f"• Total records: {total_records}\n"
                    f"• Separate sheet per folder\n"
                    f"• RA-D-PS format with dynamic columns\n\n"
                    f"📄 File: {os.path.basename(output_path)}"
                )
                
                progress_window.destroy()
                messagebox.showinfo("Export Complete", success_msg)
                open_file_cross_platform(output_path)
            else:
                progress_window.destroy()
                messagebox.showwarning("No Data", "No data was found in any of the selected folders.")
        
        except Exception as e:
            progress_window.destroy()
            messagebox.showerror("Export Error", f"Error creating combined Excel file:\n{str(e)}")
            print(f"Combined export error: {traceback.format_exc()}")

    def _process_multiple_folders(self, folders):
        """Process the selected folders and create separate Excel files for each"""
        if not folders:
            return
        
        # Process each folder and create separate Excel files
        successful_exports = 0
        total_folders = len(folders)
        
        # enhanced progress dialog with live script updates
        progress_window = tk.Toplevel(self.master)
        progress_window.title("Processing Folders - Live Progress")
        progress_window.geometry("700x500")
        progress_window.configure(bg="#d7e3fc")
        progress_window.transient(self.master)
        progress_window.grab_set()
        
        # center progress window on main window
        progress_window.update_idletasks()
        main_x = self.master.winfo_x()
        main_y = self.master.winfo_y()
        main_width = self.master.winfo_width()
        main_height = self.master.winfo_height()
        
        dialog_width = progress_window.winfo_width()
        dialog_height = progress_window.winfo_height()
        
        x = main_x + (main_width - dialog_width) // 2
        y = main_y + (main_height - dialog_height) // 2
        progress_window.geometry(f"+{x}+{y}")
        
        # main progress label
        progress_label = tk.Label(progress_window, text="Initializing...", 
                                 font=("Aptos", 12, "bold"), bg="#d7e3fc")
        progress_label.pack(pady=10)
        
        # progress bar
        progress_bar = tk.Frame(progress_window, bg="#ddd", height=20)
        progress_bar.pack(pady=5, padx=20, fill=tk.X)
        
        progress_fill = tk.Frame(progress_bar, bg="#4CAF50", height=20)
        progress_fill.pack(side=tk.LEFT, fill=tk.Y)
        
        # live activity log with scrollbar
        log_frame = tk.Frame(progress_window, bg="#d7e3fc")
        log_frame.pack(pady=10, padx=20, fill=tk.BOTH, expand=True)
        
        log_text = tk.Text(log_frame, font=("Consolas", 9), bg="#f8f8f8", fg="#333", 
                          wrap=tk.WORD, height=20, width=80)
        scrollbar = tk.Scrollbar(log_frame, orient="vertical")
        log_text.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=log_text.yview)
        
        log_text.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # status summary
        status_label = tk.Label(progress_window, text="Ready to start processing...", 
                               font=("Aptos", 10), bg="#d7e3fc", fg="#666")
        status_label.pack(pady=5)
        
        def log_message(message, level="INFO"):
            """add timestamped message to the live log"""
            timestamp = datetime.datetime.now().strftime("%H:%M:%S")
            if level == "ERROR":
                color_tag = "error"
                prefix = "❌"
            elif level == "SUCCESS":
                color_tag = "success"  
                prefix = "✅"
            elif level == "PARSING":
                color_tag = "parsing"
                prefix = "🔍"
            elif level == "FILE":
                color_tag = "file"
                prefix = "📄"
            else:
                color_tag = "info"
                prefix = "ℹ️"
            
            log_entry = f"[{timestamp}] {prefix} {message}\n"
            log_text.insert(tk.END, log_entry, color_tag)
            log_text.see(tk.END)
            
            # configure text tags for colors
            log_text.tag_configure("error", foreground="#d32f2f")
            log_text.tag_configure("success", foreground="#388e3c") 
            log_text.tag_configure("parsing", foreground="#1976d2")
            log_text.tag_configure("file", foreground="#f57c00")
            log_text.tag_configure("info", foreground="#333")
            
            progress_window.update()
        
        def update_progress(current, total, current_folder=""):
            """update progress bar and main label"""
            if total > 0:
                progress = current / total
                progress_fill.config(width=int(660 * progress))
            progress_label.config(text=f"Processing folder {current} of {total}: {current_folder}")
            status_label.config(text=f"Folders: {current}/{total} | Files processed: varies by folder")
            progress_window.update()
        
        log_message("Starting multi-folder processing session")
        log_message(f"Total folders to process: {total_folders}")
        
        for i, folder in enumerate(folders, 1):
            folder_name = os.path.basename(folder)
            update_progress(i, total_folders, folder_name)
            log_message(f"Processing folder: {folder_name}")
            
            # track files for this folder
            successfully_parsed_files = []
            failed_files = []
            empty_files = []
            
            try:
                # scan folder for xml files
                log_message("Scanning folder for XML files...", "PARSING")
                xml_files = []
                for f in os.listdir(folder):
                    if (f.lower().endswith(".xml") and 
                        not f.startswith("._") and
                        os.path.isfile(os.path.join(folder, f))):
                        xml_files.append(os.path.join(folder, f))
                
                if not xml_files:
                    log_message(f"No XML files found in folder: {folder_name}", "ERROR")
                    messagebox.showwarning("No XML Files", f"No XML files found in:\n{folder_name}")
                    continue
                
                log_message(f"Found {len(xml_files)} XML files to process", "SUCCESS")
                
                # initialize data structures for this folder
                all_data = []
                parse_cases = defaultdict(list)
                
                log_message("Starting XML parsing for individual files...", "PARSING")
                
                # parse each xml file with detailed logging
                for j, xml_file in enumerate(sorted(xml_files), 1):
                    file_name = os.path.basename(xml_file)
                    log_message(f"Parsing file {j}/{len(xml_files)}: {file_name}", "FILE")
                    
                    # update the status to show file progress within folder
                    status_label.config(text=f"Folder {i}/{total_folders} | File {j}/{len(xml_files)} | {file_name[:30]}...")
                    progress_window.update()
                    
                    try:
                        # parse this individual file
                        df, unblinded_df = parse_radiology_sample(xml_file)
                        parse_case = detect_parse_case(xml_file)
                        
                        log_message(f"  └─ Detected parse case: {parse_case}", "PARSING")
                        
                        # check if we got any data from this file
                        has_main_data = not df.empty
                        has_unblinded_data = not unblinded_df.empty
                        
                        if has_main_data or has_unblinded_data:
                            successfully_parsed_files.append(file_name)
                            
                            # process main data
                            if has_main_data:
                                data_rows = df.to_dict('records')
                                all_data.extend(data_rows)
                                parse_cases[parse_case].extend(data_rows)
                                log_message(f"  └─ Extracted {len(data_rows)} main data rows", "SUCCESS")
                            
                            # process unblinded data  
                            if has_unblinded_data:
                                unblinded_rows = unblinded_df.to_dict('records')
                                all_data.extend(unblinded_rows)
                                parse_cases[parse_case].extend(unblinded_rows)
                                log_message(f"  └─ Extracted {len(unblinded_rows)} unblinded data rows", "SUCCESS")
                        else:
                            empty_files.append(file_name)
                            log_message(f"  └─ No data extracted from {file_name}", "ERROR")
                            
                    except Exception as file_error:
                        failed_files.append(f"{file_name} ({str(file_error)})")
                        log_message(f"  └─ Error parsing {file_name}: {str(file_error)}", "ERROR")
                        print(f"Error parsing {file_name}: {str(file_error)}")
                
                # show parsing summary for this folder
                log_message(f"Folder parsing complete! Summary:", "SUCCESS")
                log_message(f"  ├─ Successfully parsed: {len(successfully_parsed_files)} files")
                log_message(f"  ├─ Empty/no data: {len(empty_files)} files")
                log_message(f"  └─ Failed to parse: {len(failed_files)} files")
                
                # list empty files if any
                if empty_files:
                    log_message("Files with no extractable data:", "ERROR")
                    for empty_file in empty_files[:10]:  # show first 10 to avoid spam
                        log_message(f"  • {empty_file}", "ERROR")
                    if len(empty_files) > 10:
                        log_message(f"  • ... and {len(empty_files)-10} more files", "ERROR")
                
                # list failed files if any
                if failed_files:
                    log_message("Files that failed to parse:", "ERROR")
                    for failed_file in failed_files[:10]:  # show first 10 to avoid spam
                        log_message(f"  • {failed_file}", "ERROR")
                    if len(failed_files) > 10:
                        log_message(f"  • ... and {len(failed_files)-10} more files", "ERROR")
                
                if not all_data:
                    log_message(f"No data could be extracted from folder: {folder_name}", "ERROR")
                    messagebox.showwarning("No Data", f"No data could be extracted from:\n{folder_name}")
                    continue
                
                log_message(f"Parsing complete! Total records extracted: {len(all_data)}", "SUCCESS")
                
                # data quality check
                log_message("Running data quality checks...", "PARSING")
                if not self._check_for_na_rows(all_data, folder_name):
                    log_message("User chose to skip this folder due to data quality issues", "ERROR")
                    continue  # user chose not to continue with this folder
                
                log_message("Data quality check passed", "SUCCESS")
                
                # Ask user for export format choice
                log_message("Asking user for export format...", "INFO")
                
                # Create export choice dialog
                choice_window = tk.Toplevel(progress_window)
                choice_window.title(f"Export Format for: {folder_name}")
                choice_window.geometry("450x250")
                choice_window.configure(bg="#d7e3fc")
                choice_window.transient(progress_window)
                choice_window.grab_set()
                
                # Center on progress window
                choice_window.update_idletasks()
                progress_x = progress_window.winfo_x()
                progress_y = progress_window.winfo_y()
                progress_width = progress_window.winfo_width()
                progress_height = progress_window.winfo_height()
                
                dialog_width = choice_window.winfo_width()
                dialog_height = choice_window.winfo_height()
                
                x = progress_x + (progress_width - dialog_width) // 2
                y = progress_y + (progress_height - dialog_height) // 2
                choice_window.geometry(f"+{x}+{y}")
                
                font = ("Aptos", 11, "normal")
                export_choice = {'format': None}
                
                tk.Label(choice_window, text=f"Choose export format for folder:", font=("Aptos", 12, "bold"), bg="#d7e3fc").pack(pady=10)
                tk.Label(choice_window, text=folder_name, font=("Aptos", 11, "italic"), bg="#d7e3fc", fg="#666").pack(pady=5)
                
                def choose_excel():
                    export_choice['format'] = 'excel'
                    choice_window.destroy()
                
                def choose_sqlite():
                    export_choice['format'] = 'sqlite'
                    choice_window.destroy()
                
                def choose_both():
                    export_choice['format'] = 'both'
                    choice_window.destroy()
                
                # Buttons frame
                btn_frame = tk.Frame(choice_window, bg="#d7e3fc")
                btn_frame.pack(pady=20)
                
                tk.Button(btn_frame, text="Excel Only", command=choose_excel, 
                         font=font, bg="#d7e3fc", fg="black", width=15).pack(pady=5)
                
                if SQLITE_AVAILABLE:
                    tk.Button(btn_frame, text="SQLite Only", command=choose_sqlite,
                             font=font, bg="#4CAF50", fg="black", width=15).pack(pady=5)
                    tk.Button(btn_frame, text="Both Formats", command=choose_both,
                             font=font, bg="#2196F3", fg="black", width=15).pack(pady=5)
                
                tk.Label(choice_window, text="Excel = Quick viewing | SQLite = Advanced analysis", 
                        font=("Aptos", 9, "normal"), bg="#d7e3fc", fg="#666").pack(pady=10)
                
                # Wait for user choice
                choice_window.wait_window()
                
                if not export_choice['format']:
                    log_message("User cancelled export - skipping folder", "INFO")
                    continue
                
                log_message(f"User chose: {export_choice['format']}", "INFO")
                
                # Handle Excel export
                if export_choice['format'] in ['excel', 'both']:
                    log_message("Preparing Excel export...", "PARSING")
                    default_filename = f"{folder_name}_XML_Export.xlsx"
                    
                    excel_path = filedialog.asksaveasfilename(
                        title=f"Save Excel file for folder: {folder_name}",
                        defaultextension=".xlsx",
                        filetypes=[("Excel Files", "*.xlsx")],
                        initialfile=default_filename
                    )
                    
                    if excel_path:
                        log_message(f"Exporting to RA-D-PS Excel: {os.path.basename(excel_path)}", "FILE")
                        status_label.config(text=f"Folder {i}/{total_folders} | Exporting RA-D-PS Excel...")
                        progress_window.update()
                        
                        try:
                            # Convert to RA-D-PS format
                            log_message("Converting data to RA-D-PS format...", "PROCESSING")
                            combined_dataframes = {}
                            for case, case_data_list in parse_cases.items():
                                if case_data_list:
                                    combined_dataframes[case] = pd.DataFrame(case_data_list)
                            
                            ra_d_ps_records = convert_parsed_data_to_ra_d_ps_format(combined_dataframes)
                            
                            if ra_d_ps_records:
                                # Use the folder itself as the output directory for auto-naming
                                output_folder = os.path.dirname(excel_path)
                                actual_output_path = export_excel(ra_d_ps_records, output_folder, sheet=f"{folder_name}_data")
                                log_message(f"Successfully exported RA-D-PS Excel: {os.path.basename(actual_output_path)}", "SUCCESS")
                            else:
                                log_message("No RA-D-PS records generated, skipping export", "WARNING")
                        except Exception as export_error:
                            log_message(f"RA-D-PS Excel export failed: {str(export_error)}", "ERROR")
                            messagebox.showerror("Export Error", f"Failed to export RA-D-PS Excel for {folder_name}:\n{str(export_error)}")
                    else:
                        log_message("User cancelled Excel save", "INFO")
                
                # Handle SQLite export
                if export_choice['format'] in ['sqlite', 'both'] and SQLITE_AVAILABLE:
                    log_message("Preparing SQLite export...", "PARSING")
                    default_db_filename = f"{folder_name}_analysis.db"
                    
                    db_path = filedialog.asksaveasfilename(
                        title=f"Save SQLite Database for folder: {folder_name}",
                        defaultextension=".db", 
                        filetypes=[("SQLite Database", "*.db"), ("All Files", "*.*")],
                        initialfile=default_db_filename
                    )
                    
                    if db_path:
                        log_message(f"Creating SQLite database: {os.path.basename(db_path)}", "FILE")
                        status_label.config(text=f"Folder {i}/{total_folders} | Creating database...")
                        progress_window.update()
                        
                        try:
                            with RadiologyDatabase(db_path) as db:
                                batch_id = db.insert_batch_data(all_data)
                                log_message(f"Database batch inserted: {batch_id}", "SUCCESS")
                                
                                # Create analysis Excel from database
                                excel_analysis_path = db_path.replace('.db', '_analysis.xlsx')
                                export_msg = db.export_to_excel(excel_analysis_path)
                                log_message(f"Analysis Excel exported: {os.path.basename(excel_analysis_path)}", "SUCCESS")
                                
                            log_message(f"Successfully created SQLite database: {folder_name}", "SUCCESS")
                        except Exception as db_error:
                            log_message(f"SQLite export failed: {str(db_error)}", "ERROR") 
                            messagebox.showerror("Database Error", f"Failed to create database for {folder_name}:\n{str(db_error)}")
                    else:
                        log_message("User cancelled database save", "INFO")
                    default_db_filename = f"{folder_name}_analysis.db"
                    
                    db_path = filedialog.asksaveasfilename(
                        title=f"Save SQLite database for folder: {folder_name}",
                        defaultextension=".db", 
                        filetypes=[("SQLite Database", "*.db"), ("All Files", "*.*")],
                        initialfile=default_db_filename
                    )
                    
                    if db_path:
                        log_message(f"Creating SQLite database: {os.path.basename(db_path)}", "FILE")
                        status_label.config(text=f"Folder {i}/{total_folders} | Creating database...")
                        progress_window.update()
                        
                        try:
                            with RadiologyDatabase(db_path) as db:
                                batch_id = db.insert_batch_data(all_data)
                                
                                # Generate Excel analysis alongside database
                                excel_analysis_path = db_path.replace('.db', '_analysis.xlsx')
                                db.export_to_excel(excel_analysis_path)
                                
                                log_message(f"SQLite database created: {batch_id}", "SUCCESS")
                                log_message(f"Analysis Excel generated: {os.path.basename(excel_analysis_path)}", "SUCCESS")
                        except Exception as db_error:
                            log_message(f"SQLite export failed: {str(db_error)}", "ERROR")
                            messagebox.showerror("Database Error", f"Failed to create database for {folder_name}:\n{str(db_error)}")
                    else:
                        log_message("User cancelled SQLite save", "INFO")
                
                successful_exports += 1
                
            except PermissionError as perm_error:
                log_message(f"Permission denied accessing folder: {folder_name}", "ERROR")
                messagebox.showerror("Permission Error", f"Cannot access folder:\n{folder_name}")
            except Exception as folder_error:
                log_message(f"Unexpected error processing folder: {str(folder_error)}", "ERROR")
                messagebox.showerror("Error", f"Error processing folder {folder_name}:\n{str(folder_error)}")
        
        # processing complete
        log_message("All folders processed!", "SUCCESS")
        log_message(f"Final results: {successful_exports}/{total_folders} folders exported successfully")
        
        # keep window open for a moment so user can review the log
        final_button = tk.Button(progress_window, text="Close (Processing Complete)", 
                                command=progress_window.destroy, font=("Aptos", 11, "bold"), 
                                bg="#4CAF50", fg="black", width=30)
        final_button.pack(pady=10)
        
        # auto-close after showing final summary 
        def close_with_summary():
            progress_window.destroy()
            messagebox.showinfo("Export Complete", 
                               f"Successfully exported {successful_exports} out of {total_folders} folders.\n\nCheck the live log for detailed information about each step.")
        
        progress_window.after(2000, close_with_summary)  # auto-close after 2 seconds
        
        # clear the file list since we processed folders directly
        self.files = []
        self._update_file_list()

    def select_excel(self):
        """select an existing excel file to append parsed data to"""
        path = filedialog.askopenfilename(
            title="Select Excel file to append",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if path:
            self.excel_path = path  # store selected excel file path
            messagebox.showinfo("Excel Selected", f"Will append to:\n{path}")

    def _update_file_list(self):
        """update the listbox display with the names of selected xml files"""
        self.listbox.delete(0, tk.END)  # clear existing entries
        for path in self.files:
            self.listbox.insert(tk.END, os.path.basename(path))  # show only filenames for clarity

    def parse_files(self):
        """
        parse selected files and append data to existing excel file or create new one
        includes comprehensive data validation and user feedback
        """
        if not self.files:
            messagebox.showinfo("No files", "Please select XML files to parse.")
            return
        
        case_data, case_unblinded_data = parse_multiple(self.files)
        if not case_data and not case_unblinded_data:
            messagebox.showinfo("Result", "No data parsed.")
            return

        # Check for N/A rows across all data
        all_parsed_data = []
        for df in case_data.values():
            all_parsed_data.extend(df.to_dict('records'))
        for df in case_unblinded_data.values():
            all_parsed_data.extend(df.to_dict('records'))
        
        if all_parsed_data:
            self._check_for_na_rows(all_parsed_data, "selected files")

        import pandas as pd
        import os

        # Determine output path
        if self.excel_path and os.path.exists(self.excel_path):
            output_path = self.excel_path
        else:
            output_path = os.path.join(os.path.expanduser("~"), "Desktop", "parsed_radiology_output.xlsx")

        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write main data sheets by parse case
            for case, df in case_data.items():
                sheet_name = self._sanitize_sheet_name(case, "_Main")
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Write unblinded data sheets by parse case
            for case, df in case_unblinded_data.items():
                sheet_name = self._sanitize_sheet_name(case, "_Unblinded")
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Apply formatting to all sheets
        self._format_excel_sheets_by_case(output_path, case_data, case_unblinded_data)

        total_main_rows = sum(len(df) for df in case_data.values())
        total_unblinded_rows = sum(len(df) for df in case_unblinded_data.values())
        
        messagebox.showinfo("Success", f"Parsed {len(self.files)} files into {len(case_data)} main cases and {len(case_unblinded_data)} unblinded cases.\nTotal: {total_main_rows} main rows, {total_unblinded_rows} unblinded rows\nExported to: {output_path}")
        open_file_cross_platform(output_path)

        self.master.title("NYT XML Parser")  # Reset title

    def export_to_sqlite(self):
        """Export parsed data to SQLite database with analysis capabilities"""
        if not SQLITE_AVAILABLE:
            messagebox.showerror("SQLite Unavailable", "SQLite database features are not available.\nPlease check that radiology_database.py is present.")
            return
            
        if not self.files:
            messagebox.showinfo("No files", "Please select XML files to parse.")
            return
        
        # Ask user where to save the database file
        db_path = filedialog.asksaveasfilename(
            title="Save SQLite Database",
            defaultextension=".db",
            filetypes=[("SQLite Database", "*.db"), ("All Files", "*.*")],
            initialfile="radiology_analysis.db"
        )
        if not db_path:
            return

        try:
            # Parse the XML files
            self.master.title("NYT XML Parser - Parsing files...")
            case_data, case_unblinded_data = parse_multiple(self.files)
            
            if not case_data and not case_unblinded_data:
                messagebox.showinfo("Result", "No data parsed.")
                return

            # Combine all data for database insertion
            all_parsed_data = []
            for df in case_data.values():
                all_parsed_data.extend(df.to_dict('records'))
            for df in case_unblinded_data.values():
                unblinded_records = df.to_dict('records')
                # Mark unblinded data
                for record in unblinded_records:
                    record['is_unblinded'] = True
                all_parsed_data.extend(unblinded_records)

            if not all_parsed_data:
                messagebox.showinfo("No Data", "No data could be extracted from the selected files.")
                return

            # Check data quality
            if not self._check_for_na_rows(all_parsed_data, "selected files"):
                return

            self.master.title("NYT XML Parser - Creating database...")
            
            # Create database and insert data
            with RadiologyDatabase(db_path) as db:
                batch_id = db.insert_batch_data(all_parsed_data)
                
                # Generate analysis report
                self.master.title("NYT XML Parser - Generating analysis...")
                quality_report = db.get_quality_report()
                
                # Create Excel export alongside database
                excel_path = db_path.replace('.db', '_analysis.xlsx')
                export_msg = db.export_to_excel(excel_path)

            # Show comprehensive success message
            stats = quality_report['overall_stats']
            message_parts = [
                f"✅ SQLite Database created successfully!",
                f"\n📊 DATABASE SUMMARY:",
                f"• Database: {os.path.basename(db_path)}",
                f"• Excel Export: {os.path.basename(excel_path)}",
                f"• Total Files: {stats.get('total_files', 0)}",
                f"• Total Nodules: {stats.get('total_nodules', 0)}",
                f"• Total Ratings: {stats.get('total_ratings', 0)}",
                f"• Batch ID: {batch_id}",
                f"\n🔍 ANALYSIS FEATURES:",
                f"• Nodule-centric relational structure",
                f"• Radiologist agreement analysis", 
                f"• Data quality tracking",
                f"• SQL query capabilities",
                f"• Excel export for compatibility"
            ]

            # Add quality issues if any
            if quality_report['quality_issues']:
                message_parts.extend([
                    f"\n⚠️ QUALITY ISSUES DETECTED:",
                    f"• Check 'Quality Issues' sheet in Excel"
                ])
                for issue in quality_report['quality_issues'][:3]:  # Show first 3
                    message_parts.append(f"• {issue['issue_type']}: {issue['issue_count']} instances")

            messagebox.showinfo("SQLite Export Complete", "\n".join(message_parts))
            
            # Ask user if they want to open the results
            if messagebox.askyesno("Open Results", "Would you like to open the Excel analysis file?"):
                open_file_cross_platform(excel_path)

            self.master.title("NYT XML Parser")

        except Exception as e:
            self.master.title("NYT XML Parser")
            error_msg = f"Error creating SQLite database:\n{str(e)}"
            messagebox.showerror("Database Error", error_msg)
            print(f"SQLite export error: {traceback.format_exc()}")

    def export_new_excel(self):
        if not self.files:
            messagebox.showinfo("No files", "Please select XML files to parse.")
            return
        
        case_data, case_unblinded_data = parse_multiple(self.files)
        if not case_data and not case_unblinded_data:
            messagebox.showinfo("Result", "No data parsed.")
            return

        # Check for N/A rows across all data
        all_parsed_data = []
        for df in case_data.values():
            all_parsed_data.extend(df.to_dict('records'))
        for df in case_unblinded_data.values():
            all_parsed_data.extend(df.to_dict('records'))
        
        if all_parsed_data:
            if not self._check_for_na_rows(all_parsed_data, "selected files"):
                return  # User chose not to continue

        # Ask user where to save the new excel file
        path = filedialog.asksaveasfilename(
            title="Save Excel file",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if not path:
            return

        try:
            # Transform data to match your template format
            print("🔄 Transforming to template format...")
            
            # Combine all data 
            combined_data = []
            for df in case_data.values():
                combined_data.extend(df.to_dict('records'))
            for df in case_unblinded_data.values():
                combined_data.extend(df.to_dict('records'))
            
            # Transform to repeating Radiologist 1-4 format
            template_data = self._transform_to_template_format(combined_data)
            
            # Create Excel with template formatting
            self._create_template_excel(template_data, path)
            
            messagebox.showinfo("Success", f"✅ Excel file created with template format!\n\n📊 Summary:\n• Total data rows: {len(template_data)}\n• Radiologist columns: 1, 2, 3, 4 (repeating)\n• Color-coded by radiologist\n• Auto-fitted columns\n\nExported to: {os.path.basename(path)}")
            open_file_cross_platform(path)

        except Exception as e:
            messagebox.showerror("Export Error", f"Error creating Excel file:\n{str(e)}")
            print(f"Excel export error: {traceback.format_exc()}")

        self.master.title("NYT XML Parser")  # Reset title

    def _transform_to_template_format(self, all_data):
        """
        Transform data to match the template format with repeating Radiologist 1-4 columns
        """
        template_rows = []
        
        # Group data by file and nodule to organize sessions
        file_nodule_groups = {}
        
        for row in all_data:
            file_id = row.get('FileID', 'unknown')
            nodule_id = row.get('NoduleID', 'unknown')
            key = f"{file_id}_{nodule_id}"
            
            if key not in file_nodule_groups:
                file_nodule_groups[key] = []
            file_nodule_groups[key].append(row)
        
        # Transform each group into template format
        for group_key, sessions in file_nodule_groups.items():
            # Sort sessions by radiologist for consistent ordering
            sessions.sort(key=lambda x: x.get('Radiologist', ''))
            
            # Create rows with repeating Radiologist 1-4 pattern
            for i, session in enumerate(sessions):
                # Determine which radiologist column (1-4, cycling)
                rad_num = (i % 4) + 1
                
                # Create template row
                template_row = {
                    'FileID': session.get('FileID', ''),
                    'NoduleID': session.get('NoduleID', ''),
                    'ParseCase': session.get('ParseCase', ''),
                    'SessionType': session.get('SessionType', ''),
                    'SOP_UID': session.get('SOP_UID', ''),
                    'StudyInstanceUID': session.get('StudyInstanceUID', ''),
                    'SeriesInstanceUID': session.get('SeriesInstanceUID', ''),
                    'Modality': session.get('Modality', ''),
                    'DateService': session.get('DateService', ''),
                    'TimeService': session.get('TimeService', ''),
                    'X_coord': session.get('X_coord', ''),
                    'Y_coord': session.get('Y_coord', ''),
                    'Z_coord': session.get('Z_coord', ''),
                    'CoordCount': session.get('CoordCount', ''),
                }
                
                # Add radiologist data to appropriate column (1-4)
                for j in range(1, 5):
                    if j == rad_num:
                        # This radiologist's data
                        confidence = session.get('Confidence', '')
                        subtlety = session.get('Subtlety', '')
                        obscuration = session.get('Obscuration', '')
                        reason = session.get('Reason', '')
                        
                        # Create compact rating string
                        ratings = []
                        if confidence not in ['', '#N/A', 'MISSING']:
                            ratings.append(f"Conf:{confidence}")
                        if subtlety not in ['', '#N/A', 'MISSING']:
                            ratings.append(f"Sub:{subtlety}")
                        if obscuration not in ['', '#N/A', 'MISSING']:
                            ratings.append(f"Obs:{obscuration}")
                        if reason not in ['', '#N/A', 'MISSING']:
                            ratings.append(f"Reason:{reason}")
                        
                        template_row[f'Radiologist {j}'] = " | ".join(ratings) if ratings else ""
                    else:
                        # Empty column for other radiologists
                        template_row[f'Radiologist {j}'] = ""
                
                # Add actual radiologist ID for reference
                template_row['ActualRadiologist'] = session.get('Radiologist', '')
                template_row['RadiologistSlot'] = rad_num
                
                template_rows.append(template_row)
        
        return template_rows

    def _create_template_excel(self, template_data, excel_path):
        """
        Create Excel file with template formatting matching your design
        """
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Radiology Analysis"
        
        # Define column headers to match your template
        headers = [
            'FileID', 'NoduleID', 'ParseCase', 'SessionType',
            'Radiologist 1', 'Radiologist 2', 'Radiologist 3', 'Radiologist 4',
            'SOP_UID', 'StudyInstanceUID', 'SeriesInstanceUID', 
            'X_coord', 'Y_coord', 'Z_coord', 'CoordCount',
            'Modality', 'DateService', 'TimeService'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Define radiologist colors (light hues for alternating pattern)
        rad_colors = {
            1: "E3F2FD",  # Light Blue
            2: "E8F5E9",  # Light Green  
            3: "FFF3E0",  # Light Orange
            4: "F3E5F5"   # Light Purple
        }
        
        # Write data rows with alternating colors
        row_num = 2
        for data_row in template_data:
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col)
                cell.value = data_row.get(header, "")
                
                # Apply alternating color pattern
                if header.startswith('Radiologist '):
                    # Get radiologist number
                    rad_num = int(header.split()[-1])
                    
                    # Alternate white and colored columns
                    if col % 2 == 0:  # Even columns get color
                        if data_row.get(header, "").strip():  # Only color if has data
                            cell.fill = PatternFill(start_color=rad_colors[rad_num], 
                                                   end_color=rad_colors[rad_num], 
                                                   fill_type="solid")
                    # Odd columns stay white (default)
                else:
                    # For non-radiologist columns, light alternating pattern
                    if col % 2 == 0:
                        cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                
                # Highlight MISSING values in orange
                if str(cell.value) == "MISSING":
                    cell.fill = PatternFill(start_color="FFE0B3", end_color="FFE0B3", fill_type="solid")
                
                # Center alignment for better readability
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            row_num += 1
        
        # Auto-fit all columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            
            # Set optimal width (with some padding)
            adjusted_width = min(max_length + 3, 50)  # Cap at 50 for very long content
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders for better visual separation
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
        
        # Save the workbook
        wb.save(excel_path)
        print(f"✅ Template Excel created: {excel_path}")

    def export_ra_d_ps_excel(self):
        """Export data using the new RA-D-PS format with auto-naming and versioning"""
        if not self.files:
            messagebox.showinfo("No files", "Please select XML files to parse.")
            return
        
        # Determine output folder based on how files were selected
        folder_path = None
        
        # Check if all files are from the same folder (folder selection mode)
        if len(self.files) > 1:
            common_folder = os.path.dirname(self.files[0])
            if all(os.path.dirname(f) == common_folder for f in self.files):
                # All files from same folder, use that folder for output
                folder_path = common_folder
        
        # If not from folder selection or mixed folders, ask user
        if folder_path is None:
            folder_path = filedialog.askdirectory(
                title="Select folder to save RA-D-PS Excel file"
            )
            if not folder_path:
                return

        try:
            self.master.title("NYT XML Parser - Parsing files...")
            
            # Parse the files
            case_data, case_unblinded_data = parse_multiple(self.files)
            if not case_data and not case_unblinded_data:
                messagebox.showinfo("Result", "No data parsed.")
                return

            # Convert parsed data to RA-D-PS format
            print("🔄 Converting to RA-D-PS format...")
            
            # Properly combine main and unblinded data per parse case
            combined_case_data = {}
            all_cases = set(case_data.keys()) | set(case_unblinded_data.keys())
            
            print(f"📋 Found parse cases: {list(all_cases)}")
            
            for case in all_cases:
                main_df = case_data.get(case, pd.DataFrame())
                unblinded_df = case_unblinded_data.get(case, pd.DataFrame())
                
                print(f"  📊 Case '{case}': main={len(main_df)} rows, unblinded={len(unblinded_df)} rows")
                
                # Combine as tuple for RA-D-PS conversion
                combined_case_data[case] = (main_df, unblinded_df)
            
            # Convert each case to RA-D-PS format
            all_ra_d_ps_records = []
            for case, (main_df, unblinded_df) in combined_case_data.items():
                print(f"🔄 Converting case '{case}' to RA-D-PS...")
                case_records = convert_parsed_data_to_ra_d_ps_format((main_df, unblinded_df))
                all_ra_d_ps_records.extend(case_records)
                print(f"  ✅ Generated {len(case_records)} records for case '{case}'")
            
            ra_d_ps_records = all_ra_d_ps_records
            
            if not ra_d_ps_records:
                messagebox.showinfo("Result", "No data to export in RA-D-PS format.")
                return

            # Export to Excel with auto-naming
            print("📊 Exporting to RA-D-PS Excel format...")
            output_path = export_excel(ra_d_ps_records, folder_path, sheet="radiology_data")
            
            # Calculate summary statistics
            total_records = len(ra_d_ps_records)
            total_radiologists = sum(len(rec.get("radiologists", {})) for rec in ra_d_ps_records)
            max_radiologists = max((len(rec.get("radiologists", {})) for rec in ra_d_ps_records), default=0)
            
            success_msg = (
                f"✅ RA-D-PS Excel export completed!\n\n"
                f"📊 Summary:\n"
                f"• Total records: {total_records}\n"
                f"• Total radiologist entries: {total_radiologists}\n"
                f"• Max radiologists per record: {max_radiologists}\n"
                f"• Auto-named with timestamp\n"
                f"• Alternating row striping\n"
                f"• Spacer columns for clarity\n\n"
                f"📄 File: {os.path.basename(output_path)}"
            )
            
            messagebox.showinfo("Export Complete", success_msg)
            open_file_cross_platform(output_path)

        except Exception as e:
            messagebox.showerror("Export Error", f"Error creating RA-D-PS Excel file:\n{str(e)}")
            print(f"RA-D-PS export error: {traceback.format_exc()}")

        self.master.title("NYT XML Parser")  # Reset title

    def _sanitize_sheet_name(self, case_name, suffix=""):
        """
        Create a valid Excel sheet name from case name and suffix
        
        Args:
            case_name: The original case name
            suffix: Optional suffix to add (like "_Main" or "_Unblinded")
            
        Returns:
            str: Valid Excel sheet name (max 31 chars, no invalid characters)
        """
        # Remove invalid characters for Excel sheet names
        invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
        sanitized = case_name
        for char in invalid_chars:
            sanitized = sanitized.replace(char, '_')
        
        # Add suffix
        full_name = sanitized + suffix
        
        # Excel sheet names must be 31 characters or less
        if len(full_name) > 31:
            # Truncate case name to make room for suffix
            max_case_len = 31 - len(suffix)
            sanitized = sanitized[:max_case_len]
            full_name = sanitized + suffix
        
        return full_name

    def _format_excel_sheets_by_case(self, output_path, case_data, case_unblinded_data):
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        from openpyxl.utils import get_column_letter

        wb = load_workbook(output_path)
        
        # Define colors for alternating files
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        blue_fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
        
        # Format main data sheets
        for case, df in case_data.items():
            sheet_name = f"Main_{case}"[:31]
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                self._apply_alternating_colors(ws, df, white_fill, blue_fill)
                self._auto_fit_columns(ws)
        
        # Format unblinded data sheets
        for case, df in case_unblinded_data.items():
            sheet_name = f"Unblinded_{case}"[:31]
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                self._apply_alternating_colors(ws, df, white_fill, blue_fill)
                self._auto_fit_columns(ws)
        
        wb.save(output_path)

    def _apply_alternating_colors(self, worksheet, df, white_fill, blue_fill):
        if df.empty:
            return
        
        from openpyxl.styles import PatternFill
        
        # Define base colors for different parse cases (light, medium pairs)
        case_base_colors = {
            "Complete_Attributes": ["E8F5E8", "D0F0D0"],      # Light & Medium Green
            "With_Reason_Partial": ["FFF8DC", "F0E68C"],      # Light & Medium Yellow
            "Core_Attributes_Only": ["E6F2FF", "B3D9FF"],     # Light & Medium Blue
            "Minimal_Attributes": ["FFE4E1", "FFB6C1"],       # Light & Medium Pink
            "No_Characteristics": ["F5F5F5", "E0E0E0"],       # Light & Medium Gray
            "LIDC_Single_Session": ["F0E6FF", "E0CCFF"],      # Light & Medium Purple
            "LIDC_Multi_Session_2": ["E0FFFF", "B0FFFF"],     # Light & Medium Cyan
            "LIDC_Multi_Session_3": ["F5FFFA", "E0FFF0"],     # Light & Medium Mint
            "LIDC_Multi_Session_4": ["FDF5E6", "F5DEB3"],     # Light & Medium Wheat
            "Unknown_Structure": ["FFCCCB", "FF9999"],        # Light & Medium Red
            "Parse_Error": ["FF6B6B", "FF4444"],              # Medium & Dark Red
            # Add new cases
            "No_Sessions_Found": ["FFE4B5", "FFD700"],        # Light & Medium Orange
            "No_Reads_Found": ["DDA0DD", "DA70D6"],           # Light & Medium Orchid
            "XML_Parse_Error": ["CD5C5C", "B22222"],          # Indian Red & Fire Brick
            "Detection_Error": ["F0F8FF", "E6E6FA"],          # Alice Blue & Lavender
        }
        # Default colors and special fills
        default_colors = ["FFFFFF", "F8F8F8"]  # White & Very Light Gray
        missing_fill = PatternFill(start_color="FFE0B3", end_color="FFE0B3", fill_type="solid")  # Light Orange for MISSING
        
        # Get unique FileIDs to determine alternation
        unique_files = df['FileID'].unique()
        file_to_color_index = {file_id: idx % 2 for idx, file_id in enumerate(unique_files)}
        
        # Apply colors row by row
        for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
            parse_case = row.get('ParseCase', 'Unknown')
            file_id = row['FileID']
            
            # Get colors for this parse case
            colors = case_base_colors.get(parse_case, default_colors)
            
            # Determine which color to use (alternating by FileID)
            color_index = file_to_color_index[file_id]
            base_color_hex = colors[color_index]
            
            # Create base fill for this row
            base_fill = PatternFill(start_color=base_color_hex, end_color=base_color_hex, fill_type="solid")
            
            # Apply colors to each cell in the row
            for col_idx, (col_name, cell_value) in enumerate(row.items(), start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                # Check if value is MISSING and highlight in orange
                if str(cell_value) == "MISSING":
                    cell.fill = missing_fill
                else:
                    cell.fill = base_fill

    def _auto_fit_columns(self, worksheet):
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column
            for cell in col:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            adjusted_width = max_length + 2
            worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width

    def _add_hyperlinks(self, worksheet, df):
        # Add hyperlinks from FileID column in Unblinded Reads to Main Data sheet
        file_id_col = None
        for col_idx, cell in enumerate(worksheet[1], 1):
            if cell.value == "FileID":
                file_id_col = col_idx
                break
        
        if file_id_col:
            for row_idx in range(2, len(df) + 2):
                cell = worksheet.cell(row=row_idx, column=file_id_col)
                # Create hyperlink to Main Data sheet
                cell.hyperlink = f"#'Main Data'!A1"
                cell.style = "Hyperlink"

    def _sanitize_sheet_name(self, case_name, suffix=""):
        """Ensure sheet name is valid for Excel"""
        # Remove invalid characters and truncate
        sanitized = re.sub(r'[\\/*?:\[\]]', '_', case_name)
        full_name = f"{sanitized}{suffix}"
        return full_name[:31]  # Excel sheet name limit

    def _is_blank_row_fast(self, row_data):
        """Fast check if a row is blank (all empty values)"""
        try:
            return all(
                str(value).strip() == "" or value is None or 
                (hasattr(pd, 'isna') and pd.isna(value))
                for value in row_data
            )
        except Exception:
            return False

    def clear_files(self):
        """clear the selected file list and provide user feedback"""
        self.files = []
        self._update_file_list()
        self.show_temporary_error("File list cleared.")

    def show_creator_signature(self):
        # Bring main window to front first
        self.master.lift()
        self.master.focus_force()

        popup = tk.Toplevel(self.master)
        popup.overrideredirect(True)
        popup.configure(bg="#0A1628")  # Dark blue background like the logo
        popup.attributes("-topmost", True)

        self.master.update_idletasks()
        # DIMENSIONS: 350px wide x 55px height (reduced from 350x80)
        width = 350
        height = 55
        win_x = self.master.winfo_x()
        win_y = self.master.winfo_y()
        win_w = self.master.winfo_width()
        x = win_x + (win_w // 2) - (width // 2)
        y_start = win_y - height
        y_end = win_y + 15  # Reduced slide distance (was 20, now 15)

        popup.geometry(f"{width}x{height}+{x}+{y_start}")

        # Main container frame with minimal padding for 55px height
        main_frame = tk.Frame(popup, bg="#0A1628")
        main_frame.pack(expand=True, fill=tk.BOTH, padx=6, pady=2)  # Further reduced padding

        # Logo frame
        logo_frame = tk.Frame(main_frame, bg="#0A1628")
        logo_frame.pack(side=tk.TOP, fill=tk.X)

        # ASCII Art Logo (smaller font)
        logo_text = ">_[ISA:SYS]"
        logo_label = tk.Label(
            logo_frame,
            text=logo_text,
            bg="#0A1628",
            fg="#00D4FF",  # Bright cyan like the logo
            font=("Courier New", 12, "bold"),  # Reduced from 14 to 12
            justify=tk.CENTER
        )
        logo_label.pack(pady=(2, 0))  # Minimal padding

        # Creator info frame
        info_frame = tk.Frame(main_frame, bg="#0A1628")
        info_frame.pack(side=tk.TOP, fill=tk.X)

        # Creator text only (smaller font)
        creator_label = tk.Label(
            info_frame,
            text="Created by: Isa Lucia Schlichting",
            bg="#0A1628",
            fg="#FFFFFF",
            font=("Aptos", 9, "bold")  # Reduced from 10 to 9
        )
        creator_label.pack(pady=(0, 2))  # Minimal padding

        # Bring popup to front and focus
        popup.lift()
        popup.focus_force()
        popup.attributes("-topmost", True)
        popup.after_idle(popup.lift)

        # Slide down animation
        steps = 20         # smoother
        duration = 600     # slower (was 200)
        delay = duration // steps
        delta = (y_end - y_start) / steps

        def slide_down(step=0):
            new_y = int(y_start + delta * step)
            popup.geometry(f"{width}x{height}+{x}+{new_y}")
            if step < steps:
                popup.after(delay, slide_down, step + 1)
            else:
                popup.after(3500, slide_up, steps)  # Show for 3.5 seconds

        def slide_up(step):
            new_y = int(y_start + delta * step)
            popup.geometry(f"{width}x{height}+{x}+{new_y}")
            if step > 0:
                popup.after(delay, slide_up, step - 1)
            else:
                popup.destroy()

        slide_down()

    def show_temporary_error(self, message):
        popup = tk.Toplevel(self.master)
        popup.overrideredirect(True)
        popup.configure(bg="#FF6B6B")  # red
        popup.attributes("-topmost", True)

        self.master.update_idletasks()
        width = max(300, len(message) * 8)
        height = 60
        win_x = self.master.winfo_x()
        win_y = self.master.winfo_y()
        win_w = self.master.winfo_width()
        win_h = self.master.winfo_height()
        x = win_x + (win_w // 2) - (width // 2)
        y = win_y + win_h + 10

        popup.geometry(f"{width}x{height}+{x}+{y}")

        label = tk.Label(
            popup,
            text=message,
            bg="#FF6B6B",
            fg="white",
            font=("Aptos", 10, "bold"),
            wraplength=width - 20
        )
        label.pack(expand=True, fill=tk.BOTH)

        popup.lift()
        popup.focus_force()
        popup.attributes("-topmost", True)
        popup.after_idle(popup.lift)

        popup.after(2500, popup.destroy)

    def _check_for_na_rows(self, all_data, folder_name):
        """
        check for rows with all n/a values and warn the user about data quality issues
        
        args:
            all_data: list of dictionaries containing parsed data
            folder_name: name of folder/source for user feedback
            
        returns:
            bool: true if user wants to continue, false to cancel
        """
        if not all_data:
            return True
        
        na_rows = 0
        total_rows = len(all_data)
        
        # Define columns to check (excluding metadata that might legitimately be N/A)
        important_columns = ['Confidence', 'Subtlety', 'Obscuration', 'Reason', 'X_coord', 'Y_coord', 'SOP_UID']
        
        for row in all_data:
            # Check if all important values are N/A
            na_count = sum(1 for col in important_columns if row.get(col) == "#N/A")
            if na_count == len(important_columns):
                na_rows += 1
        
        if na_rows > 0:
            percentage = (na_rows / total_rows) * 100
            warning_msg = (
                f"Warning: Found {na_rows} out of {total_rows} rows ({percentage:.1f}%) "
                f"with all N/A values in folder '{folder_name}'.\n\n"
                f"This might indicate:\n"
                f"• Empty or malformed XML files\n"
                f"• XML structure not matching expected format\n"
                f"• Missing data in source files\n\n"
                f"Do you want to continue with the export?"
            )
            
            if not messagebox.askyesno("Data Quality Warning", warning_msg):
                return False
        
        return True

    def _add_blank_rows_between_files(self, all_data):
        """
        Add blank separator rows between different files for better visual separation
        
        Args:
            all_data: list of data dictionaries
            
        Returns:
            list: processed data with blank rows inserted
        """
        if not all_data:
            return []
            
        processed_data = []
        current_file = None
        
        for row in all_data:
            file_id = row.get('FileID', '')
            
            # Add blank row when file changes (except for first file)
            if current_file is not None and current_file != file_id:
                # Add blank separator row
                blank_row = {key: '' for key in row.keys()}
                blank_row['FileID'] = '--- FILE SEPARATOR ---'
                processed_data.append(blank_row)
            
            processed_data.append(row)
            current_file = file_id
            
        return processed_data

    def _add_file_separators_preserve_nodules(self, all_data):
        """
        Add file separator rows while preserving nodule groupings for Standard Sessions
        
        This ensures that all radiologists for each nodule stay grouped together,
        with file separators only between different files (not between nodules).
        
        Args:
            all_data: list of data dictionaries sorted by FileID, NoduleID, Radiologist
            
        Returns:
            list: processed data with file separators that preserve nodule grouping
        """
        if not all_data:
            return []
            
        processed_data = []
        current_file = None
        
        for row in all_data:
            file_id = row.get('FileID', '')
            
            # Add blank row when file changes (except for first file)
            if current_file is not None and current_file != file_id:
                # Add blank separator row
                blank_row = {key: '' for key in row.keys()}
                blank_row['FileID'] = '--- FILE SEPARATOR ---'
                processed_data.append(blank_row)
            
            processed_data.append(row)
            current_file = file_id
            
        return processed_data

    def _format_sheet(self, worksheet, df, case_colors, sheet_type, original_data):
        """
        Apply comprehensive formatting to Excel worksheet
        
        Args:
            worksheet: openpyxl worksheet object
            df: pandas DataFrame with the data
            case_colors: dict mapping parse cases to colors
            sheet_type: string describing the sheet type
            original_data: original data list for additional context
        """
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        if df.empty:
            return
            
        # Define special fills
        missing_fill = PatternFill(start_color="FFE0B3", end_color="FFE0B3", fill_type="solid")  # Light orange
        separator_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light gray
        
        # Format header row
        for col_idx, cell in enumerate(worksheet[1], 1):
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply formatting to data rows
        for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
            parse_case = row.get('ParseCase', 'Unknown')
            file_id = row.get('FileID', '')
            
            # Handle separator rows
            if file_id == '--- FILE SEPARATOR ---':
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = separator_fill
                    cell.font = Font(italic=True, color="666666")
                continue
            
            # Get base color for this parse case
            base_color = case_colors.get(parse_case, "FFFFFF")
            base_fill = PatternFill(start_color=base_color, end_color=base_color, fill_type="solid")
            
            # Apply colors to each cell in the row
            for col_idx, (col_name, cell_value) in enumerate(row.items(), start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                # Highlight MISSING values in orange
                if str(cell_value) == "MISSING":
                    cell.fill = missing_fill
                    cell.font = Font(color="CC0000")  # Dark red text
                else:
                    cell.fill = base_fill
                
                # Center alignment for better readability
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-fit columns
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            
            # Set optimal width (with some padding)
            adjusted_width = min(max_length + 3, 50)  # Cap at 50 for very long content
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders for better visual separation
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border

    def _format_standard_sessions_sheet(self, worksheet, df, case_colors, sheet_type, original_data):
        """
        Apply specialized formatting to Standard Sessions sheet with enhanced visual organization
        
        Features:
        - Alternating column colors (base color vs white)
        - File sections alternate between light green and light purple
        - Proper sorting by FileID -> NoduleID -> Radiologist
        
        Args:
            worksheet: openpyxl worksheet object
            df: pandas DataFrame with the data
            case_colors: dict mapping parse cases to colors
            sheet_type: string describing the sheet type
            original_data: original data list for additional context
        """
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        if df.empty:
            return
            
        # Define alternating file section colors
        file_colors = {
            'primary': "E6FFE6",    # Light green
            'secondary': "F0E6FF"   # Light purple
        }
        
        # Define special fills
        missing_fill = PatternFill(start_color="FFE0B3", end_color="FFE0B3", fill_type="solid")  # Light orange
        separator_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light gray
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White
        
        # Format header row
        for col_idx, cell in enumerate(worksheet[1], 1):
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Track current file for alternating section colors
        current_file_id = None
        file_section_index = 0
        
        # Apply formatting to data rows
        for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
            parse_case = row.get('ParseCase', 'Unknown')
            file_id = row.get('FileID', '')
            
            # Handle separator rows
            if file_id == '--- FILE SEPARATOR ---':
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = separator_fill
                    cell.font = Font(italic=True, color="666666")
                continue
            
            # Track file changes for section coloring
            if current_file_id != file_id and file_id != '':
                current_file_id = file_id
                file_section_index += 1
            
            # Determine base color for this file section
            section_color_key = 'primary' if file_section_index % 2 == 1 else 'secondary'
            base_section_color = file_colors[section_color_key]
            
            # Create fill objects for alternating columns
            base_fill = PatternFill(start_color=base_section_color, end_color=base_section_color, fill_type="solid")
            
            # Apply colors to each cell in the row with alternating column pattern
            for col_idx, (col_name, cell_value) in enumerate(row.items(), start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                # Highlight MISSING values in orange (priority over other colors)
                if str(cell_value) == "MISSING":
                    cell.fill = missing_fill
                    cell.font = Font(color="CC0000")  # Dark red text
                else:
                    # Alternate between base section color and white for columns
                    if col_idx % 2 == 1:  # Odd columns get section color
                        cell.fill = base_fill
                    else:  # Even columns get white
                        cell.fill = white_fill
                
                # Center alignment for better readability
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-fit columns
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            
            # Set optimal width (with some padding)
            adjusted_width = min(max_length + 3, 50)  # Cap at 50 for very long content
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders for better visual separation
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border

    def _export_with_formatting_detailed(self, all_data, parse_cases, excel_path, log_message=None):
        """
        Enhanced export method specifically for multi-folder processing with live logging
        
        Args:
            all_data: list of all data dictionaries
            parse_cases: dict organizing data by parse case
            excel_path: output Excel file path
            log_message: optional logging function for progress updates
        """
        if log_message:
            log_message("Starting detailed Excel export with formatting...")
        
        # Use the existing _export_with_formatting method but with logging support
        self._export_with_formatting(all_data, parse_cases, excel_path)
        
        if log_message:
            log_message("Detailed Excel export completed successfully")

    def _export_with_formatting(self, all_data, parse_cases, excel_path):
        """
        export data to excel with comprehensive formatting, color coding, and organized sheets
        
        features:
        - alternating row colors based on parse case
        - separate sheets for detailed vs standard coordinate sessions  
        - missing value highlighting in orange
        - auto-hide columns with >85% n/a values
        - blank separator rows between different files
        
        args:
            all_data: list of all parsed data records (original radiologist-per-row format)
            parse_cases: dictionary organizing data by parse case
            excel_path: output file path for the excel file
        """
        try:
            # Use original data format (radiologist per row) instead of nodule-centric
            print("📊 Preparing original format data...")
            
            # Process data to add blank rows between files
            processed_data = self._add_blank_rows_between_files(all_data)
            
            # Create DataFrame with processed data
            df = pd.DataFrame(processed_data)
            
            # Separate data by coordinate complexity (using original logic)
            detailed_data = []
            standard_data = []
            
            for row in processed_data:
                # Safely get CoordCount and convert to int for comparison
                coord_count = row.get('CoordCount', 0)
                try:
                    coord_count_int = int(coord_count) if coord_count is not None else 0
                except (ValueError, TypeError):
                    coord_count_int = 0
                
                if row.get('SessionType') == 'Detailed' or coord_count_int > 10:
                    detailed_data.append(row)
                else:
                    standard_data.append(row)
            
            # Define colors for each parse case
            case_colors = {
                "Complete_Attributes": "E6F3FF",        # Light blue
                "With_Reason_Partial": "FFE6E6",        # Light red  
                "Core_Attributes_Only": "E6FFE6",       # Light green
                "Minimal_Attributes": "FFFACD",         # Light yellow
                "No_Characteristics": "F0E6FF",         # Light purple
                "LIDC_Single_Session": "F5DEB3",        # Light wheat
                "LIDC_Multi_Session_2": "E0FFFF",       # Light cyan
                "LIDC_Multi_Session_3": "F0FFF0",       # Light honeydew
                "LIDC_Multi_Session_4": "E6E6FA",       # Light lavender (for detailed sessions)
                "No_Sessions_Found": "FFE4B5",          # Light moccasin
                "No_Reads_Found": "DDA0DD",             # Light plum
                "XML_Parse_Error": "FF6B6B",            # Light red
                "Detection_Error": "F0F8FF",            # Alice blue
                "Unknown": "F5F5F5"                     # Light gray
            }

            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                # Write main sheet with all data
                if len(df) > 0:
                    df.to_excel(writer, sheet_name='All Data', index=False)
                    self._format_sheet(writer.sheets['All Data'], df, case_colors, "General", all_data)
                
                # Create separate sheet for detailed coordinate sessions
                if detailed_data:
                    detailed_df = pd.DataFrame(detailed_data)
                    detailed_df.to_excel(writer, sheet_name='Detailed Coordinates', index=False)
                    self._format_sheet(writer.sheets['Detailed Coordinates'], detailed_df, case_colors, "Detailed", detailed_data)
                    
                    print(f"Created 'Detailed Coordinates' sheet with {len(detailed_data)} rows containing extensive coordinate data")
                
                # Create sheet for standard sessions
                if standard_data:
                    # Sort standard data properly: FileID -> NoduleID -> Radiologist  
                    sorted_standard_data = sorted(standard_data, key=lambda x: (
                        x.get('FileID', ''),
                        int(x.get('NoduleID', 0)) if str(x.get('NoduleID', 0)).isdigit() else 999999,
                        x.get('Radiologist', '')
                    ))
                    
                    # Process with file separators
                    processed_standard_data = self._add_blank_rows_between_files(sorted_standard_data)
                    standard_df = pd.DataFrame(processed_standard_data)
                    standard_df.to_excel(writer, sheet_name='Standard Sessions', index=False)
                    self._format_sheet(writer.sheets['Standard Sessions'], standard_df, case_colors, "Standard", processed_standard_data)
                
                # Create separate sheets for each parse case
                for case, case_data in parse_cases.items():
                    if case_data:
                        processed_case_data = self._add_blank_rows_between_files(case_data)
                        case_df = pd.DataFrame(processed_case_data)
                        sheet_name = f"Parse {case}" if case.startswith("Case") else case
                        # Truncate sheet name if too long
                        if len(sheet_name) > 31:
                            sheet_name = sheet_name[:31]
                        
                        case_df.to_excel(writer, sheet_name=sheet_name, index=False)
                        self._format_sheet(writer.sheets[sheet_name], case_df, case_colors, case, case_data)

            # Count and report different session types with safe type handling
            detailed_count = len(detailed_data)
            standard_count = len(standard_data)
            
            # Count MISSING values safely with better type handling
            missing_count = 0
            for row in all_data:
                for key, value in row.items():
                    try:
                        # Handle None, empty, and string values safely
                        if value is None:
                            continue  # None is not MISSING, it's just null
                        str_value = str(value)
                        if str_value == "MISSING":
                            missing_count += 1
                    except Exception:
                        continue  # Skip values that can't be converted to string
            
            total_values = len(all_data) * len(df.columns) if len(df) > 0 and len(all_data) > 0 else 1
            missing_percentage = (missing_count / total_values * 100) if total_values > 0 else 0.0
            
            # Enhanced success message (original format)
            message_parts = [
                f"Data exported successfully to:\n{excel_path}\n",
                f"📊 SESSION BREAKDOWN:",
                f"• All Data: {len(all_data)} total rows (radiologist sessions)",
                f"• Detailed Coordinates: {detailed_count} rows (>10 coordinates each)",
                f"• Standard Sessions: {standard_count} rows",
                f"• Parse case sheets: {len(parse_cases)} sheets"
            ]
            
            if missing_count > 0:
                message_parts.extend([
                    f"\n⚠️ DATA QUALITY:",
                    f"• {missing_count} MISSING values ({missing_percentage:.1f}% of data)",
                    f"• MISSING values highlighted in light orange",
                    f"• Legend: MISSING = Expected but not found, N/A = Not expected"
                ])
            else:
                message_parts.append(f"\n✅ No MISSING values detected!")
            
            messagebox.showinfo("Export Complete", "\n".join(message_parts))
            
        except TypeError as te:
            error_msg = f"Type Error in Excel export:\n{str(te)}\n\nThis usually means there's a data type mismatch. Check your data for mixed types."
            messagebox.showerror("Export Error", error_msg)
            print(f"Detailed TypeError: {traceback.format_exc()}")
        except Exception as e:
            error_msg = f"Error exporting to Excel:\n{str(e)}"
            messagebox.showerror("Export Error", error_msg)
            print(f"Detailed error: {traceback.format_exc()}")

    def _format_standard_sessions_sheet_template(self, worksheet, df, case_colors, sheet_type, original_data):
        """
        Apply formatting to Standard Sessions sheet matching the user's template exactly
        
        Features:
        - Alternating light blue and light green columns as shown in template
        - Proper grouping with FileID/NoduleID repeating for each radiologist
        - Clean visual separation
        
        Args:
            worksheet: openpyxl worksheet object
            df: pandas DataFrame with the data
            case_colors: dict mapping parse cases to colors (not used in this template)
            sheet_type: string describing the sheet type
            original_data: original data list for additional context
        """
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        if df.empty:
            return
            
        # Define template colors matching the user's image
        light_blue = "ADD8E6"      # Light blue for alternating columns
        light_green = "90EE90"     # Light green for alternating columns  
        header_blue = "4472C4"     # Dark blue for header
        
        # Create fill objects
        blue_fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type="solid")
        green_fill = PatternFill(start_color=light_green, end_color=light_green, fill_type="solid")
        header_fill = PatternFill(start_color=header_blue, end_color=header_blue, fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Format header row
        for col_idx, cell in enumerate(worksheet[1], 1):
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply alternating column colors to data rows
        for row_idx in range(2, len(df) + 2):  # Start from row 2 (after header)
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                # Skip empty rows (our spacing rows)
                if all(str(worksheet.cell(row=row_idx, column=c).value).strip() == '' 
                      for c in range(1, len(df.columns) + 1)):
                    continue
                
                # Alternate between light blue and light green for columns
                # Following the pattern in the user's image
                if col_idx in [1, 3, 5, 7, 9, 11, 13]:  # Odd columns - light blue
                    cell.fill = blue_fill
                elif col_idx in [2, 4, 6, 8, 10, 12, 14]:  # Even columns - light green
                    cell.fill = green_fill
                else:
                    cell.fill = white_fill
                
                # Center alignment for better readability
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-fit columns
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            
            # Set optimal width (with some padding)
            adjusted_width = min(max_length + 3, 50)  # Cap at 50 for very long content
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders for better visual separation
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border


def detect_parse_case(file_path):
    """
    Detect the structure/case of an XML file for appropriate parsing strategy
    """
    try:
        tree = ET.parse(file_path)
        root = tree.getroot()
        
        # Get namespace if present
        m = re.match(r'\{(.*)\}', root.tag)
        ns_uri = m.group(1) if m else ''
        
        def tag(name):
            return f"{{{ns_uri}}}{name}" if ns_uri else name
        
        # Check for basic structure indicators
        header = root.find(tag('ResponseHeader'))
        sessions = root.findall(tag('readingSession')) or root.findall(tag('CXRreadingSession'))
        
        if not sessions:
            return "No_Sessions_Found"
        
        # Analyze first session for characteristics
        first_session = sessions[0]
        unblinded_reads = first_session.findall(tag('unblindedReadNodule')) or first_session.findall(tag('unblindedRead'))
        
        if not unblinded_reads:
            return "No_Reads_Found"
        
        # Analyze first read for characteristics
        first_read = unblinded_reads[0]
        characteristics = first_read.find(tag('characteristics'))
        
        if characteristics is None:
            return "No_Characteristics"
        
        # Count available characteristics
        char_fields = ['confidence', 'subtlety', 'obscuration', 'reason']
        available_chars = []
        for field in char_fields:
            elem = characteristics.find(tag(field))
            if elem is not None and elem.text:
                available_chars.append(field)
        
        # Determine case based on available characteristics and header completeness
        header_complete = header is not None
        modality_present = False
        if header:
            modality_elem = header.find(tag('Modality'))
            modality_present = modality_elem is not None and modality_elem.text
        
        # Classification logic
        if len(available_chars) >= 3 and 'reason' in available_chars and header_complete and modality_present:
            return "Complete_Attributes"
        elif 'reason' in available_chars and len(available_chars) >= 2:
            return "With_Reason_Partial"
        elif len(available_chars) >= 2 and 'confidence' in available_chars and 'subtlety' in available_chars:
            return "Core_Attributes_Only"
        elif len(available_chars) == 1:
            return "Minimal_Attributes"
        elif len(sessions) == 1:
            return "LIDC_Single_Session"
        elif len(sessions) == 2:
            return "LIDC_Multi_Session_2"
        elif len(sessions) == 3:
            return "LIDC_Multi_Session_3"
        elif len(sessions) == 4:
            return "LIDC_Multi_Session_4"
        else:
            return "Unknown_Structure"
            
    except Exception as e:
        print(f"Error detecting parse case for {file_path}: {e}")
        return "XML_Parse_Error"


if __name__ == "__main__":
    # This allows the module to be run directly for testing
    root = tk.Tk()
    app = NYTXMLGuiApp(root)
    root.mainloop()
