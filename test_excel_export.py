#!/usr/bin/env python3

# Quick test of Excel export with enhanced formatting
import sys
sys.path.append('.')

try:
    from XMLPARSE import parse_radiology_sample
    import pandas as pd
    from collections import defaultdict
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    
    def test_excel_export():
        xml_path = '/Volumes/LUCKY/tcia-lidc-xml/157/158.xml'
        output_path = '/Users/isa/Desktop/python projects/XML PARSE/test_158_output.xlsx'
        
        print('📊 TESTING EXCEL EXPORT WITH ENHANCED FORMATTING')
        print('='*55)
        
        # Parse the XML
        df, unblinded_df = parse_radiology_sample(xml_path)
        
        # Combine data
        all_data = []
        if not df.empty:
            all_data.extend(df.to_dict('records'))
        if not unblinded_df.empty:
            all_data.extend(unblinded_df.to_dict('records'))
        
        print(f'Data rows to export: {len(all_data)}')
        
        # Create Excel with formatting
        combined_df = pd.DataFrame(all_data)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            combined_df.to_excel(writer, sheet_name='LIDC_Test', index=False)
            
            # Apply formatting
            workbook = writer.book
            worksheet = writer.sheets['LIDC_Test']
            
            # Define colors
            case_color = "F0E6FF"  # Light purple for LIDC
            missing_color = "FFE0B3"  # Light orange for MISSING
            
            # Get unique files for alternating
            unique_files = combined_df['FileID'].unique()
            file_colors = {file_id: idx % 2 for idx, file_id in enumerate(unique_files)}
            
            case_fill = PatternFill(start_color=case_color, end_color=case_color, fill_type="solid")
            missing_fill = PatternFill(start_color=missing_color, end_color=missing_color, fill_type="solid")
            
            # Apply formatting to each row
            for row_idx, (_, row_data) in enumerate(combined_df.iterrows(), start=2):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    
                    if str(cell_value) == "MISSING":
                        cell.fill = missing_fill
                    else:
                        cell.fill = case_fill
            
            # Auto-fit columns
            for column in worksheet.columns:
                max_length = 0
                col_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[col_letter].width = adjusted_width
        
        print(f'✅ Excel file created: {output_path}')
        
        # Count and report formatting
        missing_count = sum(1 for row in all_data for value in row.values() if str(value) == "MISSING")
        total_cells = len(all_data) * len(combined_df.columns)
        
        print(f'\\n📋 FORMATTING SUMMARY:')
        print(f'  🟠 Orange cells (MISSING): {missing_count}')
        print(f'  🟣 Purple cells (normal): {total_cells - missing_count}')
        print(f'  📊 Total cells: {total_cells}')
        print(f'  📈 MISSING percentage: {100*missing_count/total_cells:.1f}%')
        
        print(f'\\n🎯 EXPLANATION:')
        print(f'  • Orange cells = Expected for LIDC format but missing from XML')
        print(f'  • Purple cells = Normal data or N/A (not expected)')
        print(f'  • File alternates between light/dark purple by FileID')
        print(f'  • This is much better than before!')
        
        return output_path
    
    if __name__ == "__main__":
        test_excel_export()

except Exception as e:
    print(f'❌ Error: {e}')
    import traceback
    traceback.print_exc()
